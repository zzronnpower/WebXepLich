# -*- coding: utf-8 -*-
from __future__ import annotations

from datetime import date, datetime, timedelta
import json
import logging
import os
import threading
import time
from io import BytesIO
from uuid import uuid4

from fastapi import Depends, FastAPI, Form, HTTPException, Request
from fastapi.responses import JSONResponse, PlainTextResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy import func, or_, text
from sqlalchemy.orm import Session

from backend.app import crud, models
from backend.app.db import CoSo, dong_co, lay_phien_lam_viec
from backend.app.scheduler.solver import (
    danh_sach_ngay_trong_tuan,
    giai_lich_tuan,
    lay_trong_so,
    tao_nhu_cau_chich_ngoai,
)
from backend.app.services.schedule_service import chay_tu_xep_tuan, chay_xep_lich_tuan, lay_thong_ke_xep_lich
from backend.app.services.job_service import lay_job, tao_job_xep_lich
from backend.app.seed import tao_du_lieu_mau


ung_dung = FastAPI(title="Lich lam viec")
ung_dung.mount("/static", StaticFiles(directory="backend/app/static"), name="static")
giao_dien = Jinja2Templates(directory="backend/app/templates")
logger = logging.getLogger("xeplich.web")
_http_metric_lock = threading.Lock()
_http_metric: dict[str, dict[str, float | int]] = {}


def tra_template(ten_template: str, context: dict, **kwargs):
    request = context.get("request")
    if request is None:
        raise ValueError("Template context thieu request")
    return giao_dien.TemplateResponse(request=request, name=ten_template, context=context, **kwargs)


def _khoi_tao_logging() -> None:
    if logging.getLogger().handlers:
        return
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s %(message)s",
    )


def _ghi_nhan_http_metric(method: str, path: str, status_code: int, duration_ms: float) -> None:
    key = f"{method.upper()} {path}"
    with _http_metric_lock:
        item = _http_metric.get(key)
        if item is None:
            item = {"count": 0, "errors": 0, "sum_ms": 0.0, "max_ms": 0.0}
            _http_metric[key] = item
        item["count"] = int(item["count"]) + 1
        item["sum_ms"] = float(item["sum_ms"]) + duration_ms
        item["max_ms"] = max(float(item["max_ms"]), duration_ms)
        if status_code >= 400:
            item["errors"] = int(item["errors"]) + 1


def _tong_hop_http_metric() -> dict[str, dict[str, float | int]]:
    with _http_metric_lock:
        result: dict[str, dict[str, float | int]] = {}
        for key, item in _http_metric.items():
            count = int(item["count"])
            avg_ms = round(float(item["sum_ms"]) / count, 2) if count else 0.0
            result[key] = {
                "count": count,
                "errors": int(item["errors"]),
                "avg_ms": avg_ms,
                "max_ms": round(float(item["max_ms"]), 2),
            }
        return result


def _la_route_nhay_cam(path: str, method: str) -> bool:
    if method.upper() != "POST":
        return False
    return "/xoa" in path


def _doc_admin_token_tu_request(request: Request) -> str | None:
    return request.headers.get("x-admin-token") or request.query_params.get("admin_token")


@ung_dung.middleware("http")
async def middleware_request_id(request: Request, call_next):
    request_id = request.headers.get("x-request-id") or str(uuid4())
    request.state.request_id = request_id
    token_he_thong = os.getenv("ADMIN_TOKEN", "").strip()
    if token_he_thong and _la_route_nhay_cam(request.url.path, request.method):
        token_gui_len = _doc_admin_token_tu_request(request)
        if token_gui_len != token_he_thong:
            return JSONResponse(status_code=403, content={"detail": "Thiếu hoặc sai ADMIN_TOKEN cho thao tác này."})
    bat_dau = time.perf_counter()
    try:
        response = await call_next(request)
    except Exception:
        duration_ms = round((time.perf_counter() - bat_dau) * 1000.0, 2)
        _ghi_nhan_http_metric(request.method, request.url.path, 500, duration_ms)
        logger.exception(
            "http_request_error req_id=%s method=%s path=%s duration_ms=%s",
            request_id,
            request.method,
            request.url.path,
            duration_ms,
        )
        raise
    duration_ms = round((time.perf_counter() - bat_dau) * 1000.0, 2)
    _ghi_nhan_http_metric(request.method, request.url.path, response.status_code, duration_ms)
    response.headers["X-Request-ID"] = request_id
    logger.info(
        "http_request req_id=%s method=%s path=%s status=%s duration_ms=%s",
        request_id,
        request.method,
        request.url.path,
        response.status_code,
        duration_ms,
    )
    return response


@ung_dung.on_event("startup")
def khoi_tao():
    _khoi_tao_logging()
    for _ in range(10):
        try:
            CoSo.metadata.create_all(bind=dong_co)
            with next(lay_phien_lam_viec()) as phien:
                tao_du_lieu_mau(phien)
                cap_nhat_nhom_off(phien)
                dam_bao_nhom_he_thong(phien)
                dam_bao_xoa_796adv(phien)
                dam_bao_cot_thu_tu(phien)
                dam_bao_cot_ngay_nghi_nguon(phien)
                dam_bao_cot_ten_lich(phien)
                dam_bao_cot_spa_off_ghi_chu(phien)
                dam_bao_trong_so(phien)
                dam_bao_chi_muc_hieu_nang(phien)
            return
        except Exception:
            time.sleep(1)
    CoSo.metadata.create_all(bind=dong_co)
    with next(lay_phien_lam_viec()) as phien:
        tao_du_lieu_mau(phien)
        cap_nhat_nhom_off(phien)
        dam_bao_nhom_he_thong(phien)
        dam_bao_xoa_796adv(phien)
        dam_bao_cot_thu_tu(phien)
        dam_bao_cot_ngay_nghi_nguon(phien)
        dam_bao_cot_ten_lich(phien)
        dam_bao_cot_spa_off_ghi_chu(phien)
        dam_bao_trong_so(phien)
        dam_bao_chi_muc_hieu_nang(phien)


THU_TU_NHOM_MAC_DINH = {
    "326TTV": 1,
    "197LT5": 2,
    "CN": 3,
    "PHU_SPA": 4,
    "OFF": 5,
    "CHUA_XEP": 6,
}


def ten_thu(ngay: date) -> str:
    thu = ["Hai", "Ba", "Tư", "Năm", "Sáu", "Bảy", "Chủ nhật"]
    return thu[ngay.weekday()]


def ten_nhom_hien_thi(ten_nhom: str) -> str:
    if ten_nhom == "CN":
        return "Chích ngoài"
    if ten_nhom == "PHU_SPA":
        return "Phụ Spa"
    if ten_nhom in {"Spa", "OFF"}:
        return "OFF (Ngày nghỉ)"
    if ten_nhom == "CHUA_XEP":
        return "Chưa xếp"
    return ten_nhom


def ten_trong_so_hien_thi(khoa: str) -> str:
    mapping = {
        "uu_tien_ca_ua_thich": "Ưu tiên ca ưa thích",
        "phat_ca_tranh": "Phạt ca tránh",
        "cong_bang_chich_ngoai": "Công bằng chích ngoài",
        "cong_bang_cuoi_tuan": "Công bằng cuối tuần",
        "khong_di_chich_ngoai": "Không đi chích ngoài",
        "uu_tien_ca_quan_trong": "Ưu tiên ca quan trọng",
        "han_che_ca_muon_sang": "Hạn chế ca muộn/sáng",
        "bat_buoc_chich_ngoai": "Bắt buộc chích ngoài",
    }
    return mapping.get(khoa, khoa)


def dinh_dang_ngay(ngay: date) -> str:
    return ngay.strftime("%d/%m/%Y")


def lay_thu_hai_tiep_theo(ngay_hien_tai: date | None = None) -> date:
    ngay_goc = ngay_hien_tai or date.today()
    so_ngay_con_lai = (7 - ngay_goc.weekday()) % 7
    if so_ngay_con_lai == 0:
        so_ngay_con_lai = 7
    return ngay_goc + timedelta(days=so_ngay_con_lai)


def tao_ma_nv(phien: Session) -> str:
    max_id = phien.query(func.max(models.NhanVien.id)).scalar() or 0
    return f"BS{max_id + 1:02d}"


def parse_trong_so_form(form) -> list[int]:
    ket_qua = []
    for idx in range(1, 4):
        ts_id = form.get(f"trong_so_{idx}_id")
        if not ts_id:
            continue
        try:
            parsed = _parse_int_or_none(ts_id, f"trong_so_{idx}_id")
        except ValueError:
            continue
        if parsed is not None:
            ket_qua.append(parsed)
    return ket_qua


def _parse_int_or_none(raw_value, ten_truong: str) -> int | None:
    if raw_value in (None, ""):
        return None
    try:
        return int(str(raw_value).strip())
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Trường '{ten_truong}' phải là số nguyên hợp lệ.") from exc


def _parse_int_required(raw_value, ten_truong: str) -> int:
    parsed = _parse_int_or_none(raw_value, ten_truong)
    if parsed is None:
        raise ValueError(f"Thiếu trường bắt buộc: {ten_truong}.")
    return parsed


def _parse_date_required(raw_value, ten_truong: str = "ngay") -> date:
    if raw_value in (None, ""):
        raise ValueError(f"Thiếu trường bắt buộc: {ten_truong}.")
    try:
        return datetime.fromisoformat(str(raw_value).strip()).date()
    except ValueError as exc:
        raise ValueError(f"Trường '{ten_truong}' không đúng định dạng YYYY-MM-DD.") from exc


def kiem_tra_admin_token(request: Request, admin_token: str | None = None) -> None:
    token_he_thong = os.getenv("ADMIN_TOKEN", "").strip()
    if not token_he_thong:
        return
    token_gui_len = (
        request.headers.get("x-admin-token")
        or admin_token
        or request.query_params.get("admin_token")
    )
    if token_gui_len != token_he_thong:
        raise HTTPException(status_code=403, detail="Thiếu hoặc sai ADMIN_TOKEN cho thao tác này.")


def danh_sach_ca_theo_nhom() -> dict[str, list[str]]:
    return {
        "326TTV": ["8h-19h", "8h30-19h30", "9h-20h", "10h-21h"],
        "197LT5": ["8h-19h", "9h-20h", "10h-21h"],
        "CN": ["9h-20h"],
        "PHU_SPA": ["8h30-19h30"],
        "OFF": ["Nghỉ"],
        "CHUA_XEP": ["Chưa xếp"],
    }


def tao_ca_ids_theo_nhom(phien: Session, nhom_list: list[models.NhomHienThi]) -> dict[int, list[int]]:
    ca_theo_ten = {ca.ten_ca: ca.id for ca in phien.query(models.CaLam).all()}
    ca_theo_nhom_ten = danh_sach_ca_theo_nhom()
    ket_qua: dict[int, list[int]] = {}
    for nhom in nhom_list:
        ds_ten_ca = ca_theo_nhom_ten.get(nhom.ten_nhom, [])
        ket_qua[nhom.id] = [ca_theo_ten[ten_ca] for ten_ca in ds_ten_ca if ten_ca in ca_theo_ten]
    return ket_qua


def ca_id_theo_thu_tu(ds_ca_id: list[int], thu_tu: int) -> int | None:
    if thu_tu < 1:
        return None
    vi_tri_ca = (thu_tu - 1) // 2
    if vi_tri_ca >= len(ds_ca_id):
        return None
    return ds_ca_id[vi_tri_ca]


def tao_ca_theo_ten(phien: Session) -> dict[str, int]:
    return {ca.ten_ca: ca.id for ca in phien.query(models.CaLam).all()}


def la_nhom_off(ten_nhom: str | None) -> bool:
    if not ten_nhom:
        return False
    return ten_nhom.strip().lower() in {"spa", "off", "nghi", "ngay nghi", "off (ngay nghi)"}


def la_trang_thai_nhap(trang_thai: str | None) -> bool:
    return bool(trang_thai and trang_thai.startswith("NHAP_"))


def trang_thai_chinh_thuc_tu_nhap(trang_thai: str | None) -> str:
    if not la_trang_thai_nhap(trang_thai):
        return trang_thai or "DA_XEP"
    return str(trang_thai).replace("NHAP_", "", 1) or "DA_XEP"


def tao_trang_thai_nhap(trang_thai_goc: str) -> str:
    trang_thai_sach = (trang_thai_goc or "DA_XEP").strip() or "DA_XEP"
    if trang_thai_sach.startswith("NHAP_"):
        return trang_thai_sach
    return f"NHAP_{trang_thai_sach}"


def cap_nhat_nhom_off(phien: Session):
    nhom_off = phien.query(models.NhomHienThi).filter(models.NhomHienThi.ten_nhom == "OFF").first()
    nhom_spa = phien.query(models.NhomHienThi).filter(models.NhomHienThi.ten_nhom == "Spa").first()
    if nhom_spa and not nhom_off:
        nhom_spa.ten_nhom = "OFF"
        phien.flush()
        phien.query(models.MappingNhom).filter(models.MappingNhom.nhom_hien_thi_id == nhom_spa.id).delete()
        phien.commit()
        return
    if nhom_spa and nhom_off:
        phien.query(models.MappingNhom).filter(models.MappingNhom.nhom_hien_thi_id == nhom_spa.id).delete()
        phien.delete(nhom_spa)
        phien.commit()
        return
    if nhom_off:
        phien.query(models.MappingNhom).filter(models.MappingNhom.nhom_hien_thi_id == nhom_off.id).delete()
        phien.commit()


def dam_bao_nhom_he_thong(phien: Session):
    mac_dinh = {
        "PHU_SPA": "#f2d7ff",
        "CHUA_XEP": "#f4f4f4",
    }
    da_co = {n.ten_nhom: n for n in phien.query(models.NhomHienThi).all()}
    can_commit = False
    for ten_nhom, mau_nen in mac_dinh.items():
        if ten_nhom in da_co:
            continue
        phien.add(models.NhomHienThi(ten_nhom=ten_nhom, mau_nen=mau_nen))
        can_commit = True
    if can_commit:
        phien.commit()


def dam_bao_xoa_796adv(phien: Session):
    try:
        nhom_chua_xep = phien.query(models.NhomHienThi).filter(models.NhomHienThi.ten_nhom == "CHUA_XEP").first()
        if not nhom_chua_xep:
            nhom_chua_xep = models.NhomHienThi(ten_nhom="CHUA_XEP", mau_nen="#f4f4f4")
            phien.add(nhom_chua_xep)
            phien.flush()

        nhom_796 = phien.query(models.NhomHienThi).filter(models.NhomHienThi.ten_nhom == "796ADV").first()
        ds_chi_nhanh_796 = (
            phien.query(models.ChiNhanh)
            .filter(or_(models.ChiNhanh.ma_chi_nhanh == "796", models.ChiNhanh.ten_chi_nhanh == "796ADV"))
            .all()
        )

        ds_cn_796_id = [cn.id for cn in ds_chi_nhanh_796]
        nhom_796_id = nhom_796.id if nhom_796 else None
        if not ds_cn_796_id and not nhom_796_id:
            return

        ds_dieu_kien = []
        if ds_cn_796_id:
            ds_dieu_kien.append(models.LichChiTiet.chi_nhanh_id.in_(ds_cn_796_id))
            phien.query(models.NhuCauCa).filter(models.NhuCauCa.chi_nhanh_id.in_(ds_cn_796_id)).delete(
                synchronize_session=False
            )
            phien.query(models.MappingNhom).filter(models.MappingNhom.chi_nhanh_id.in_(ds_cn_796_id)).delete(
                synchronize_session=False
            )
            ds_nv_co_796 = (
                phien.query(models.NhanVien)
                .filter(models.NhanVien.chi_nhanh.any(models.ChiNhanh.id.in_(ds_cn_796_id)))
                .all()
            )
            for nv in ds_nv_co_796:
                nv.chi_nhanh = [cn for cn in nv.chi_nhanh if cn.id not in ds_cn_796_id]
        if nhom_796_id:
            ds_dieu_kien.append(models.LichChiTiet.nhom_hien_thi_id == nhom_796_id)
            phien.query(models.MappingNhom).filter(models.MappingNhom.nhom_hien_thi_id == nhom_796_id).delete(
                synchronize_session=False
            )

        phien.query(models.LichChiTiet).filter(or_(*ds_dieu_kien)).update(
            {
                models.LichChiTiet.nhom_hien_thi_id: nhom_chua_xep.id,
                models.LichChiTiet.chi_nhanh_id: None,
                models.LichChiTiet.ca_id: None,
            },
            synchronize_session=False,
        )

        for cn in ds_chi_nhanh_796:
            phien.delete(cn)
        if nhom_796:
            phien.delete(nhom_796)
        phien.commit()
        logger.info(
            "remove_796adv_migration done converted_rows_to_chua_xep=1 removed_branches=%s removed_group=%s",
            len(ds_chi_nhanh_796),
            bool(nhom_796_id),
        )
    except Exception:
        phien.rollback()
        logger.exception("remove_796adv_migration failed")


def dam_bao_cot_thu_tu(phien: Session):
    try:
        phien.execute(text("ALTER TABLE lich_chi_tiet ADD COLUMN IF NOT EXISTS thu_tu INTEGER DEFAULT 0"))
        phien.commit()
    except Exception:
        phien.rollback()


def dam_bao_cot_ngay_nghi_nguon(phien: Session):
    try:
        phien.execute(text("ALTER TABLE ngay_nghi ADD COLUMN IF NOT EXISTS nguon VARCHAR DEFAULT 'user'"))
        phien.execute(text("UPDATE ngay_nghi SET nguon = 'user' WHERE nguon IS NULL OR trim(nguon) = ''"))
        phien.commit()
    except Exception:
        phien.rollback()


def dam_bao_cot_ten_lich(phien: Session):
    try:
        phien.execute(text("ALTER TABLE lich_tuan ADD COLUMN IF NOT EXISTS ten_lich VARCHAR"))
        phien.execute(
            text(
                """
                UPDATE lich_tuan
                SET ten_lich = CASE
                    WHEN trang_thai = 'TU_XEP' THEN 'Tự xếp ' || to_char(ngay_bat_dau, 'DD/MM/YYYY')
                    ELSE 'Lịch tự động ' || to_char(ngay_bat_dau, 'DD/MM/YYYY')
                END
                WHERE ten_lich IS NULL OR btrim(ten_lich) = ''
                """
            )
        )
        phien.commit()
    except Exception:
        phien.rollback()


def dam_bao_cot_spa_off_ghi_chu(phien: Session):
    try:
        phien.execute(text("ALTER TABLE lich_tuan ADD COLUMN IF NOT EXISTS spa_off_ghi_chu TEXT"))
        phien.commit()
    except Exception:
        phien.rollback()


def dam_bao_trong_so(phien: Session):
    crud.tao_hoac_cap_nhat_trong_so(phien, "khong_di_chich_ngoai", 6)


def dam_bao_chi_muc_hieu_nang(phien: Session):
    cau_lenh = [
        "CREATE INDEX IF NOT EXISTS idx_ngay_nghi_ngay ON ngay_nghi (ngay)",
        "CREATE INDEX IF NOT EXISTS idx_ngay_nghi_nv_ngay ON ngay_nghi (nhan_vien_id, ngay)",
        "CREATE INDEX IF NOT EXISTS idx_nhu_cau_ca_ngay ON nhu_cau_ca (ngay)",
        "CREATE INDEX IF NOT EXISTS idx_nhu_cau_ca_ngay_cn_ca ON nhu_cau_ca (ngay, chi_nhanh_id, ca_id)",
        "CREATE INDEX IF NOT EXISTS idx_lich_chi_tiet_lich_ngay ON lich_chi_tiet (lich_tuan_id, ngay)",
        "CREATE INDEX IF NOT EXISTS idx_lich_chi_tiet_nv_ngay ON lich_chi_tiet (nhan_vien_id, ngay)",
        "CREATE INDEX IF NOT EXISTS idx_lich_tuan_trang_thai ON lich_tuan (trang_thai)",
    ]
    try:
        for sql in cau_lenh:
            phien.execute(text(sql))
        phien.commit()
    except Exception:
        phien.rollback()


def tao_context_trang_chu(
    phien: Session,
    ngay_bat_dau: date,
    *,
    loi: str | None = None,
    thieu_nhu_cau: list[str] | None = None,
    loi_ngay_nghi: str | None = None,
) -> dict:
    ngay_list_nghi, ngay_nghi_theo_ngay = tai_du_lieu_ngay_nghi(phien, ngay_bat_dau)
    context = {
        "lich": lay_lich_hien_thi(phien, None),
        "la_lich_nhap": False,
        "ds_lich_tuan": danh_sach_lich_tuan(phien),
        "danh_sach_ca_nhom": danh_sach_ca_theo_nhom(),
        "ca_theo_ten": tao_ca_theo_ten(phien),
        "ten_thu": ten_thu,
        "ten_nhom_hien_thi": ten_nhom_hien_thi,
        "ngay_mac_dinh": ngay_bat_dau,
        "ds_nhan_vien": phien.query(models.NhanVien).order_by(models.NhanVien.ten_nv).all(),
        "ngay_list_nghi": ngay_list_nghi,
        "ngay_nghi_theo_ngay": ngay_nghi_theo_ngay,
    }
    if loi:
        context["loi"] = loi
    if thieu_nhu_cau:
        context["thieu_nhu_cau"] = thieu_nhu_cau
    if loi_ngay_nghi:
        context["loi_ngay_nghi"] = loi_ngay_nghi
    return context


def tai_du_lieu_ngay_nghi(phien: Session, ngay_bat_dau: date) -> tuple[list[date], dict[date, list[models.NgayNghi]]]:
    ngay_bat_dau = ngay_bat_dau - timedelta(days=ngay_bat_dau.weekday())
    ngay_list_nghi = danh_sach_ngay_trong_tuan(ngay_bat_dau)
    ds_ngay_nghi = (
        phien.query(models.NgayNghi)
        .filter(models.NgayNghi.ngay >= ngay_list_nghi[0])
        .filter(models.NgayNghi.ngay <= ngay_list_nghi[-1])
        .filter(models.NgayNghi.trang_thai == "OFF")
        .order_by(models.NgayNghi.ngay.asc())
        .all()
    )
    ngay_nghi_theo_ngay = {ngay: [] for ngay in ngay_list_nghi}
    for nn in ds_ngay_nghi:
        if nn.ngay in ngay_nghi_theo_ngay:
            ngay_nghi_theo_ngay[nn.ngay].append(nn)
    return ngay_list_nghi, ngay_nghi_theo_ngay


def lay_lich_hien_thi(phien: Session, lich_tuan_id: int | None):
    lich_tuan = None
    if lich_tuan_id:
        lich_tuan = phien.query(models.LichTuan).filter(models.LichTuan.id == lich_tuan_id).first()
    if not lich_tuan:
        lich_tuan = (
            phien.query(models.LichTuan)
            .filter(or_(models.LichTuan.trang_thai.is_(None), ~models.LichTuan.trang_thai.like("NHAP_%")))
            .order_by(models.LichTuan.id.desc())
            .first()
        )
    if not lich_tuan:
        return None

    nhom_list = phien.query(models.NhomHienThi).all()
    nhom_list = sorted(
        nhom_list,
        key=lambda n: (THU_TU_NHOM_MAC_DINH.get(n.ten_nhom, 999), n.id),
    )
    ngay_list = danh_sach_ngay_trong_tuan(lich_tuan.ngay_bat_dau)
    ca_ids_theo_nhom = tao_ca_ids_theo_nhom(phien, nhom_list)
    ngay_nghi_set = {
        (nn.nhan_vien_id, nn.ngay)
        for nn in (
            phien.query(models.NgayNghi)
            .filter(models.NgayNghi.ngay >= ngay_list[0])
            .filter(models.NgayNghi.ngay <= ngay_list[-1])
            .filter(models.NgayNghi.trang_thai == "OFF")
            .filter(or_(models.NgayNghi.nguon == "user", models.NgayNghi.nguon.is_(None)))
            .all()
        )
    }
    lich_ct = (
        phien.query(models.LichChiTiet)
        .filter(models.LichChiTiet.lich_tuan_id == lich_tuan.id)
        .order_by(models.LichChiTiet.ngay.asc(), models.LichChiTiet.nhom_hien_thi_id.asc(), models.LichChiTiet.thu_tu.asc(), models.LichChiTiet.id.asc())
        .all()
    )
    spa_off_theo_ngay = {ngay: "" for ngay in ngay_list}
    if lich_tuan.spa_off_ghi_chu:
        try:
            spa_off_raw = json.loads(lich_tuan.spa_off_ghi_chu)
            if isinstance(spa_off_raw, dict):
                for ngay in ngay_list:
                    gia_tri = spa_off_raw.get(ngay.isoformat())
                    if isinstance(gia_tri, str):
                        spa_off_theo_ngay[ngay] = gia_tri
        except Exception:
            pass
    bang = {
        nhom.id: {ngay: [] for ngay in ngay_list}
        for nhom in nhom_list
    }
    nhom_off_id = next((n.id for n in nhom_list if la_nhom_off(n.ten_nhom)), None)
    nhom_tu_do_ids = {n.id for n in nhom_list if n.ten_nhom in {"PHU_SPA", "CHUA_XEP"}}
    for ct in lich_ct:
        if ct.nhom_hien_thi_id not in bang:
            continue
        ca_id_hien_thi = ct.ca_id
        if not ca_id_hien_thi and ct.nhom_hien_thi_id not in nhom_tu_do_ids and ct.nhom_hien_thi_id != nhom_off_id:
            ca_id_hien_thi = ca_id_theo_thu_tu(ca_ids_theo_nhom.get(ct.nhom_hien_thi_id, []), int(ct.thu_tu or 0))
        la_ngay_nghi_khoa = bool(nhom_off_id and ct.nhom_hien_thi_id == nhom_off_id and (ct.nhan_vien_id, ct.ngay) in ngay_nghi_set)
        bang[ct.nhom_hien_thi_id][ct.ngay].append(
            {
                "id": ct.id,
                "ten_nv": ct.nhan_vien.ten_nv,
                "nhan_vien_id": ct.nhan_vien_id,
                "ca_id": ca_id_hien_thi,
                "cap_do": ct.nhan_vien.cap_do,
                "thu_tu": int(ct.thu_tu or 0),
                "khoa_keo_tha": la_ngay_nghi_khoa,
            }
        )

    return {
        "lich_tuan": lich_tuan,
        "nhom_list": nhom_list,
        "ngay_list": ngay_list,
        "bang": bang,
        "spa_off_theo_ngay": spa_off_theo_ngay,
    }


def danh_sach_lich_tuan(phien: Session) -> list[models.LichTuan]:
    return (
        phien.query(models.LichTuan)
        .filter(or_(models.LichTuan.trang_thai.is_(None), ~models.LichTuan.trang_thai.like("NHAP_%")))
        .order_by(models.LichTuan.ngay_bat_dau.desc(), models.LichTuan.id.desc())
        .all()
    )


def kiem_tra_lich(phien: Session, lich_tuan_id: int | None):
    lich_tuan = None
    if lich_tuan_id:
        lich_tuan = phien.query(models.LichTuan).filter(models.LichTuan.id == lich_tuan_id).first()
    if not lich_tuan:
        lich_tuan = (
            phien.query(models.LichTuan)
            .filter(or_(models.LichTuan.trang_thai.is_(None), ~models.LichTuan.trang_thai.like("NHAP_%")))
            .order_by(models.LichTuan.id.desc())
            .first()
        )
    if not lich_tuan:
        return None

    ngay_list = danh_sach_ngay_trong_tuan(lich_tuan.ngay_bat_dau)
    ca_map = {ca.id: ca for ca in phien.query(models.CaLam).all()}
    lich_ct = (
        phien.query(models.LichChiTiet)
        .filter(models.LichChiTiet.lich_tuan_id == lich_tuan.id)
        .all()
    )

    dem_lich = {}
    for ct in lich_ct:
        key = (ct.ngay, ct.ca_id, ct.chi_nhanh_id)
        dem_lich[key] = dem_lich.get(key, 0) + 1

    errors = []
    warnings = []
    chi_nhanh_map = {cn.id: cn for cn in phien.query(models.ChiNhanh).all()}
    for key, so_nv in dem_lich.items():
        ngay, ca_id, chi_nhanh_id = key
        if not ca_id or so_nv <= 2:
            continue
        ten_chi_nhanh = chi_nhanh_map[chi_nhanh_id].ten_chi_nhanh if chi_nhanh_id in chi_nhanh_map else "Chích ngoài"
        ten_ca = ca_map[ca_id].ten_ca if ca_id in ca_map else ""
        errors.append(f"Quá 2 người/ca: {ten_chi_nhanh} {dinh_dang_ngay(ngay)} ca {ten_ca} ({so_nv} người)")

    gio_nv = {}
    ca_trong_ngay = {}
    for ct in lich_ct:
        ca = ca_map.get(ct.ca_id)
        so_gio = ca.so_gio if ca else 0
        gio_nv[ct.nhan_vien_id] = gio_nv.get(ct.nhan_vien_id, 0) + so_gio
        ngay_key = (ct.nhan_vien_id, ct.ngay)
        ca_trong_ngay[ngay_key] = ca_trong_ngay.get(ngay_key, 0) + 1

    nhan_vien_map = {nv.id: nv for nv in phien.query(models.NhanVien).all()}
    for nv_id, tong_gio in gio_nv.items():
        nv = nhan_vien_map.get(nv_id)
        if nv and tong_gio > nv.gio_toi_da_tuan:
            warnings.append(f"Vượt giờ tuần: {nv.ten_nv} ({tong_gio}/{nv.gio_toi_da_tuan} giờ)")

    for (nv_id, ngay), so_ca in ca_trong_ngay.items():
        if so_ca > 1:
            nv = nhan_vien_map.get(nv_id)
            ten = nv.ten_nv if nv else f"NV {nv_id}"
            errors.append(f"Trùng ca trong ngày: {ten} vào {dinh_dang_ngay(ngay)}")

    chi_nhanh_hop_le_theo_nv = {
        nv.id: {cn.id for cn in nv.chi_nhanh}
        for nv in nhan_vien_map.values()
    }
    da_bao_sai_cn = set()
    for ct in lich_ct:
        if not ct.chi_nhanh_id:
            continue
        nv = nhan_vien_map.get(ct.nhan_vien_id)
        if not nv:
            continue
        chi_nhanh_hop_le = chi_nhanh_hop_le_theo_nv.get(nv.id, set())
        if chi_nhanh_hop_le and ct.chi_nhanh_id not in chi_nhanh_hop_le:
            key = (ct.nhan_vien_id, ct.ngay, ct.chi_nhanh_id)
            if key in da_bao_sai_cn:
                continue
            da_bao_sai_cn.add(key)
            cn = chi_nhanh_map.get(ct.chi_nhanh_id)
            ten_cn = cn.ten_chi_nhanh if cn else ""
            errors.append(
                f"Sai chi nhánh theo hồ sơ: {nv.ten_nv} không thuộc chi nhánh {ten_cn} ({dinh_dang_ngay(ct.ngay)})."
            )

    return {
        "lich_tuan": lich_tuan,
        "errors": errors,
        "warnings": warnings,
    }


def thong_ke_phan_cong(phien: Session, phan_cong: list[dict]) -> dict[str, int]:
    if not phan_cong:
        return {
            "tong_ca": 0,
            "ca_ua_thich": 0,
            "ca_tranh": 0,
            "chich_ngoai": 0,
            "cuoi_tuan": 0,
            "ca_quan_trong": 0,
        }
    nv_map = {nv.id: nv for nv in phien.query(models.NhanVien).all()}
    dem = {
        "tong_ca": 0,
        "ca_ua_thich": 0,
        "ca_tranh": 0,
        "chich_ngoai": 0,
        "cuoi_tuan": 0,
        "ca_quan_trong": 0,
    }
    for pc in phan_cong:
        dem["tong_ca"] += 1
        nv = nv_map.get(pc["nhan_vien_id"])
        if nv:
            if pc["ca_id"] in {ca.id for ca in nv.ca_ua_thich}:
                dem["ca_ua_thich"] += 1
            if pc["ca_id"] in {ca.id for ca in nv.ca_tranh}:
                dem["ca_tranh"] += 1
        if pc.get("la_chich_ngoai"):
            dem["chich_ngoai"] += 1
        if pc["ngay"].weekday() >= 5:
            dem["cuoi_tuan"] += 1
        if pc.get("do_quan_trong", 0) > 0:
            dem["ca_quan_trong"] += 1
    return dem


@ung_dung.get("/")
def trang_chu(
    request: Request,
    lich_tuan_id: int | None = None,
    ngay_bat_dau: date | None = None,
    sap_xep: str | None = None,
    phien: Session = Depends(lay_phien_lam_viec),
):
    du_lieu_lich = lay_lich_hien_thi(phien, lich_tuan_id)
    if ngay_bat_dau:
        ngay_mac_dinh = ngay_bat_dau
    elif lich_tuan_id and du_lieu_lich:
        ngay_mac_dinh = du_lieu_lich["lich_tuan"].ngay_bat_dau
    else:
        ngay_mac_dinh = lay_thu_hai_tiep_theo()
    ds_nhan_vien = phien.query(models.NhanVien).order_by(models.NhanVien.ten_nv).all()
    ngay_list_nghi, ngay_nghi_theo_ngay = tai_du_lieu_ngay_nghi(phien, ngay_mac_dinh)
    return tra_template(
        "index.html",
        {
            "request": request,
            "lich": du_lieu_lich,
            "la_lich_nhap": bool(du_lieu_lich and la_trang_thai_nhap(du_lieu_lich["lich_tuan"].trang_thai)),
            "ds_lich_tuan": danh_sach_lich_tuan(phien),
            "danh_sach_ca_nhom": danh_sach_ca_theo_nhom(),
            "ca_theo_ten": tao_ca_theo_ten(phien),
            "ten_thu": ten_thu,
            "ten_nhom_hien_thi": ten_nhom_hien_thi,
            "ngay_mac_dinh": ngay_mac_dinh,
            "ds_nhan_vien": ds_nhan_vien,
            "ngay_list_nghi": ngay_list_nghi,
            "ngay_nghi_theo_ngay": ngay_nghi_theo_ngay,
        },
    )


@ung_dung.get("/lich-da-xep")
def lich_da_xep(request: Request, phien: Session = Depends(lay_phien_lam_viec)):
    return tra_template(
        "lich_da_xep.html",
        {
            "request": request,
            "ds_lich_tuan": danh_sach_lich_tuan(phien),
        },
    )


@ung_dung.post("/lich-da-xep/{lich_tuan_id}/doi-ten")
async def doi_ten_lich_da_xep(
    lich_tuan_id: int,
    request: Request,
    phien: Session = Depends(lay_phien_lam_viec),
):
    lich = phien.query(models.LichTuan).filter(models.LichTuan.id == lich_tuan_id).first()
    if lich:
        form = await request.form()
        ten_lich = (form.get("ten_lich") or "").strip()
        if ten_lich:
            lich.ten_lich = ten_lich
            phien.commit()
    return RedirectResponse(url="/lich-da-xep", status_code=303)


@ung_dung.post("/lich-da-xep/{lich_tuan_id}/ghi-chu")
async def cap_nhat_ghi_chu_lich_da_xep(
    lich_tuan_id: int,
    request: Request,
    phien: Session = Depends(lay_phien_lam_viec),
):
    lich = phien.query(models.LichTuan).filter(models.LichTuan.id == lich_tuan_id).first()
    if lich:
        form = await request.form()
        ghi_chu = (form.get("ghi_chu") or "").strip()
        lich.ghi_chu = ghi_chu if ghi_chu else None
        phien.commit()
    return RedirectResponse(url="/lich-da-xep", status_code=303)


@ung_dung.post("/lich-da-xep/{lich_tuan_id}/xoa")
def xoa_lich_da_xep(lich_tuan_id: int, phien: Session = Depends(lay_phien_lam_viec)):
    lich = phien.query(models.LichTuan).filter(models.LichTuan.id == lich_tuan_id).first()
    if lich:
        phien.delete(lich)
        phien.commit()
    return RedirectResponse(url="/lich-da-xep", status_code=303)


@ung_dung.get("/chatlog")
def chatlog(request: Request):
    return tra_template("chatlog.html", {"request": request})


@ung_dung.get("/healthz")
def healthz() -> dict[str, str]:
    return {"status": "ok"}


@ung_dung.get("/readyz")
def readyz(phien: Session = Depends(lay_phien_lam_viec)) -> dict[str, str]:
    try:
        phien.execute(text("SELECT 1"))
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"db_not_ready: {exc}") from exc
    return {"status": "ready"}


@ung_dung.get("/metrics")
def metrics(request: Request) -> dict[str, object]:
    return {
        "request_id": getattr(request.state, "request_id", ""),
        "http": _tong_hop_http_metric(),
        "solver": lay_thong_ke_xep_lich(),
    }


@ung_dung.get("/metrics/prometheus")
def metrics_prometheus() -> PlainTextResponse:
    lines: list[str] = []
    for key, item in _tong_hop_http_metric().items():
        metric_key = key.lower().replace(" ", "_").replace("/", "_").replace("-", "_")
        lines.append(f'http_requests_total{{route="{metric_key}"}} {item["count"]}')
        lines.append(f'http_requests_errors_total{{route="{metric_key}"}} {item["errors"]}')
        lines.append(f'http_request_avg_ms{{route="{metric_key}"}} {item["avg_ms"]}')
        lines.append(f'http_request_max_ms{{route="{metric_key}"}} {item["max_ms"]}')
    for flow, item in lay_thong_ke_xep_lich().items():
        lines.append(f'solver_runs_total{{flow="{flow}"}} {item["count"]}')
        lines.append(f'solver_runs_success_total{{flow="{flow}"}} {item["success"]}')
        lines.append(f'solver_runs_failed_total{{flow="{flow}"}} {item["failed"]}')
        lines.append(f'solver_run_avg_ms{{flow="{flow}"}} {item["avg_ms"]}')
        lines.append(f'solver_run_max_ms{{flow="{flow}"}} {item["max_ms"]}')
    return PlainTextResponse("\n".join(lines) + "\n", media_type="text/plain; version=0.0.4")


@ung_dung.post("/api/jobs/xep-lich")
async def tao_job_xep_lich_api(request: Request) -> dict[str, object]:
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    flow = str(payload.get("flow") or "xep_lich")
    ngay_bat_dau = str(payload.get("ngay_bat_dau") or "").strip()
    if not ngay_bat_dau:
        raise HTTPException(status_code=400, detail="Thieu ngay_bat_dau")
    try:
        datetime.fromisoformat(ngay_bat_dau).date()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="ngay_bat_dau khong dung dinh dang YYYY-MM-DD") from exc
    try:
        job = tao_job_xep_lich(flow=flow, ngay_bat_dau_iso=ngay_bat_dau)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return {
        "ok": True,
        "job_id": job.job_id,
        "status": job.status,
        "flow": job.flow,
        "ngay_bat_dau": job.ngay_bat_dau,
    }


@ung_dung.get("/api/jobs/{job_id}")
def xem_job_xep_lich(job_id: str) -> dict[str, object]:
    job = lay_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Khong tim thay job")
    return {
        "job_id": job.job_id,
        "flow": job.flow,
        "ngay_bat_dau": job.ngay_bat_dau,
        "status": job.status,
        "message": job.message,
        "created_at": job.created_at,
        "started_at": job.started_at,
        "finished_at": job.finished_at,
        "success": job.success,
        "lich_tuan_id": job.lich_tuan_id,
    }


@ung_dung.get("/kiem-tra")
def kiem_tra(request: Request, lich_tuan_id: int | None = None, phien: Session = Depends(lay_phien_lam_viec)):
    ket_qua = kiem_tra_lich(phien, lich_tuan_id)
    if not ket_qua:
        return RedirectResponse(url="/", status_code=303)
    return tra_template(
        "kiem_tra.html",
        {
            "request": request,
            "ket_qua": ket_qua,
        },
    )


@ung_dung.post("/cap-nhat-lich")
async def cap_nhat_lich(request: Request, phien: Session = Depends(lay_phien_lam_viec)):
    try:
        payload = await request.json()
        lich_tuan_id = payload.get("lich_tuan_id")
        khoa_nv_off = bool(payload.get("khoa_nv_off", True))
        changes = payload.get("changes", [])
        spa_off_notes = payload.get("spa_off_notes") or {}
        if not lich_tuan_id:
            return {"ok": False, "message": "Không có thay đổi."}

        lich_tuan = phien.query(models.LichTuan).filter(models.LichTuan.id == lich_tuan_id).first()
        if not lich_tuan:
            return {"ok": False, "message": "Không tìm thấy lịch tuần."}
        if not changes and not spa_off_notes:
            if la_trang_thai_nhap(lich_tuan.trang_thai):
                lich_tuan.trang_thai = trang_thai_chinh_thuc_tu_nhap(lich_tuan.trang_thai)
                phien.commit()
                return {"ok": True, "errors": [], "warnings": []}
            return {"ok": False, "message": "Không có thay đổi."}

        lich_ct = (
            phien.query(models.LichChiTiet)
            .filter(models.LichChiTiet.lich_tuan_id == lich_tuan_id)
            .all()
        )
        nhom_list = phien.query(models.NhomHienThi).all()
        nhom_map = {nhom.id: nhom for nhom in nhom_list}
        off_nhom_ids = {nhom.id for nhom in nhom_list if la_nhom_off(nhom.ten_nhom)}
        nhom_tu_do_ids = {nhom.id for nhom in nhom_list if nhom.ten_nhom in {"PHU_SPA", "CHUA_XEP"}}
        ca_ids_theo_nhom = tao_ca_ids_theo_nhom(phien, nhom_list)
        mapping_list = phien.query(models.MappingNhom).all()
        mapping_map = {(m.nhom_hien_thi_id, m.ca_id): m.chi_nhanh_id for m in mapping_list}
        ngay_set = {ct.ngay for ct in lich_ct}
        ngay_nghi_khoa = {
            (nn.nhan_vien_id, nn.ngay)
            for nn in (
                phien.query(models.NgayNghi)
                .filter(models.NgayNghi.ngay.in_(ngay_set))
                .filter(models.NgayNghi.trang_thai == "OFF")
                .filter(or_(models.NgayNghi.nguon == "user", models.NgayNghi.nguon.is_(None)))
                .all()
            )
        }
        lich_map = {ct.id: ct for ct in lich_ct}
        du_kien = {}
        off_khoa_ids = set()
        for ct in lich_ct:
            if ct.nhom_hien_thi_id in off_nhom_ids and (ct.nhan_vien_id, ct.ngay) in ngay_nghi_khoa:
                off_khoa_ids.add(ct.id)
            du_kien[ct.id] = {
                "ngay": ct.ngay,
                "nhom_hien_thi_id": ct.nhom_hien_thi_id,
                "nhan_vien_id": ct.nhan_vien_id,
                "chi_nhanh_id": ct.chi_nhanh_id,
                "ca_id": ct.ca_id,
                "thu_tu": int(ct.thu_tu or 0),
            }

        for item in changes:
            lich_id = item.get("id")
            ngay = item.get("ngay")
            nhom_id = item.get("nhom_hien_thi_id")
            if not lich_id or not ngay or not nhom_id or lich_id not in du_kien:
                continue
            lich_id = int(lich_id)
            ngay_moi = datetime.fromisoformat(ngay).date()
            thu_tu_moi = int(item.get("thu_tu") or du_kien[lich_id]["thu_tu"] or 0)
            if ngay_moi != du_kien[lich_id]["ngay"]:
                return {"ok": False, "message": "Chỉ cho phép kéo dọc trong cùng ngày."}
            if khoa_nv_off and lich_id in off_khoa_ids:
                if int(nhom_id) != int(du_kien[lich_id]["nhom_hien_thi_id"]):
                    return {"ok": False, "message": "Nhân viên OFF đã đăng ký nghỉ là khóa cứng, không thể kéo thả."}
            nhom_id = int(nhom_id)
            du_kien[lich_id]["nhom_hien_thi_id"] = nhom_id
            du_kien[lich_id]["thu_tu"] = thu_tu_moi
            if nhom_id in off_nhom_ids or nhom_id in nhom_tu_do_ids:
                du_kien[lich_id]["chi_nhanh_id"] = None
                du_kien[lich_id]["ca_id"] = None
                continue
            ca_id_payload = item.get("ca_id")
            ca_id = int(ca_id_payload) if ca_id_payload else ca_id_theo_thu_tu(ca_ids_theo_nhom.get(nhom_id, []), thu_tu_moi)
            if ca_id and ca_id not in set(ca_ids_theo_nhom.get(nhom_id, [])):
                nhom_obj = nhom_map.get(nhom_id)
                nhom_ten = nhom_obj.ten_nhom if nhom_obj else ""
                return {
                    "ok": False,
                    "message": f"Ca không hợp lệ cho nhóm {nhom_ten}.",
                }
            if not ca_id:
                nhom_obj = nhom_map.get(nhom_id)
                nhom_ten = nhom_obj.ten_nhom if nhom_obj else ""
                return {
                    "ok": False,
                    "message": f"Vị trí dòng {thu_tu_moi} vượt giới hạn 2 người/ca cho nhóm {nhom_ten}.",
                }
            chi_nhanh_id = mapping_map.get((nhom_id, ca_id))
            if (nhom_id, ca_id) not in mapping_map:
                nhom_obj = nhom_map.get(nhom_id)
                nhom_ten = nhom_obj.ten_nhom if nhom_obj else ""
                return {
                    "ok": False,
                    "message": f"Không có mapping chi nhánh cho nhóm {nhom_ten}.",
                }
            du_kien[lich_id]["ca_id"] = ca_id
            du_kien[lich_id]["chi_nhanh_id"] = chi_nhanh_id

        for lich_id, info in du_kien.items():
            nhom_id = int(info.get("nhom_hien_thi_id") or 0)
            thu_tu = int(info.get("thu_tu") or 0)
            if nhom_id in off_nhom_ids or nhom_id in nhom_tu_do_ids:
                info["chi_nhanh_id"] = None
                info["ca_id"] = None
                continue
            ca_id = int(info.get("ca_id") or 0) or ca_id_theo_thu_tu(ca_ids_theo_nhom.get(nhom_id, []), thu_tu)
            if not ca_id:
                nhom_obj = nhom_map.get(nhom_id)
                nhom_ten = nhom_obj.ten_nhom if nhom_obj else ""
                return {
                    "ok": False,
                    "message": f"Vị trí dòng {thu_tu} vượt giới hạn 2 người/ca cho nhóm {nhom_ten}.",
                }
            if ca_id not in set(ca_ids_theo_nhom.get(nhom_id, [])):
                nhom_obj = nhom_map.get(nhom_id)
                nhom_ten = nhom_obj.ten_nhom if nhom_obj else ""
                return {
                    "ok": False,
                    "message": f"Ca không hợp lệ cho nhóm {nhom_ten}.",
                }
            chi_nhanh_id = mapping_map.get((nhom_id, ca_id))
            if (nhom_id, ca_id) not in mapping_map:
                nhom_obj = nhom_map.get(nhom_id)
                nhom_ten = nhom_obj.ten_nhom if nhom_obj else ""
                return {
                    "ok": False,
                    "message": f"Không có mapping chi nhánh cho nhóm {nhom_ten}.",
                }
            info["ca_id"] = ca_id
            info["chi_nhanh_id"] = chi_nhanh_id

        dem_so_nguoi_ca = {}
        ca_map = {ca.id: ca for ca in phien.query(models.CaLam).all()}
        for info in du_kien.values():
            nhom_id = int(info.get("nhom_hien_thi_id") or 0)
            ca_id = info.get("ca_id")
            if nhom_id in off_nhom_ids or nhom_id in nhom_tu_do_ids or not ca_id:
                continue
            key = (info.get("ngay"), info.get("chi_nhanh_id"), ca_id)
            dem_so_nguoi_ca[key] = dem_so_nguoi_ca.get(key, 0) + 1
            if dem_so_nguoi_ca[key] > 2:
                nhom_obj = nhom_map.get(nhom_id)
                ten_nhom = ten_nhom_hien_thi(nhom_obj.ten_nhom) if nhom_obj else ""
                ten_ca = ca_map[ca_id].ten_ca if ca_id in ca_map else ""
                return {
                    "ok": False,
                    "message": f"Ca {ten_ca} của nhóm {ten_nhom} ngày {dinh_dang_ngay(info.get('ngay'))} chỉ tối đa 2 người.",
                }

        dem_trung = {}
        for info in du_kien.values():
            key = (info["nhan_vien_id"], info["ngay"])
            dem_trung[key] = dem_trung.get(key, 0) + 1
        bi_trung = [key for key, so_ca in dem_trung.items() if so_ca > 1]
        if bi_trung:
            nv_map = {nv.id: nv for nv in phien.query(models.NhanVien).all()}
            chi_nhanh_map = {cn.id: cn for cn in phien.query(models.ChiNhanh).all()}
            danh_sach = []
            for nv_id, ngay in bi_trung:
                nv = nv_map.get(nv_id)
                ten = nv.ten_nv if nv else nv_id
                chi_tiet = set()
                for info in du_kien.values():
                    if info["nhan_vien_id"] == nv_id and info["ngay"] == ngay:
                        cn = chi_nhanh_map.get(info["chi_nhanh_id"])
                        ca = ca_map.get(info["ca_id"])
                        ten_cn = cn.ten_chi_nhanh if cn else "Chích ngoài"
                        ten_ca = ca.ten_ca if ca else ""
                        chi_tiet.add(f"{ten_cn} {ten_ca}".strip())
                danh_sach.append(f"{ten} {dinh_dang_ngay(ngay)} ({'; '.join(sorted(chi_tiet))})")
            danh_sach = ", ".join(danh_sach)
            return {"ok": False, "message": f"Trùng ca trong ngày: {danh_sach}."}

        nv_map = {nv.id: nv for nv in phien.query(models.NhanVien).all()}
        chi_nhanh_map = {cn.id: cn for cn in phien.query(models.ChiNhanh).all()}
        canh_bao_sai_cn = []
        for info in du_kien.values():
            nv = nv_map.get(info["nhan_vien_id"])
            if not nv or not nv.chi_nhanh or not info["chi_nhanh_id"]:
                continue
            if info["chi_nhanh_id"] not in {cn.id for cn in nv.chi_nhanh}:
                cn = chi_nhanh_map.get(info["chi_nhanh_id"])
                ten_cn = cn.ten_chi_nhanh if cn else ""
                canh_bao_sai_cn.append(
                    f"Sai chi nhánh theo hồ sơ: {nv.ten_nv} không thuộc chi nhánh {ten_cn} ({dinh_dang_ngay(info['ngay'])})."
                )

        off_sau = {
            (info["nhan_vien_id"], info["ngay"])
            for info in du_kien.values()
            if info["nhom_hien_thi_id"] in off_nhom_ids
        }

        for lich_id, info in du_kien.items():
            ct = lich_map.get(lich_id)
            if not ct:
                continue
            ct.ngay = info["ngay"]
            ct.nhom_hien_thi_id = info["nhom_hien_thi_id"]
            ct.chi_nhanh_id = info["chi_nhanh_id"]
            ct.ca_id = info["ca_id"]
            ct.thu_tu = int(info["thu_tu"] or 0)

        if isinstance(spa_off_notes, dict):
            ngay_hop_le = set(danh_sach_ngay_trong_tuan(lich_tuan.ngay_bat_dau))
            spa_off_sach = {}
            for ngay_raw, noi_dung in spa_off_notes.items():
                if not isinstance(ngay_raw, str):
                    continue
                try:
                    ngay_obj = datetime.fromisoformat(ngay_raw).date()
                except ValueError:
                    continue
                if ngay_obj not in ngay_hop_le:
                    continue
                text_note = noi_dung if isinstance(noi_dung, str) else ""
                if text_note.strip():
                    spa_off_sach[ngay_obj.isoformat()] = text_note
            lich_tuan.spa_off_ghi_chu = json.dumps(spa_off_sach, ensure_ascii=False) if spa_off_sach else None

        ngay_nghi_hien_tai = (
            phien.query(models.NgayNghi)
            .filter(models.NgayNghi.ngay.in_(ngay_set))
            .filter(models.NgayNghi.trang_thai == "OFF")
            .filter(or_(models.NgayNghi.nguon == "user", models.NgayNghi.nguon.is_(None)))
            .all()
        )
        ngay_nghi_map = {(nn.nhan_vien_id, nn.ngay): nn for nn in ngay_nghi_hien_tai}
        off_khoa_set = {
            (info["nhan_vien_id"], info["ngay"])
            for info in du_kien.values()
            if info["nhom_hien_thi_id"] in off_nhom_ids and (info["nhan_vien_id"], info["ngay"]) in ngay_nghi_map
        }
        if khoa_nv_off and not off_khoa_set.issubset(off_sau):
            return {"ok": False, "message": "OFF đăng ký nghỉ là cố định, không thể kéo ra ngoài."}
        lich_tuan.trang_thai = trang_thai_chinh_thuc_tu_nhap(lich_tuan.trang_thai)
        phien.commit()
        kiem_tra = kiem_tra_lich(phien, lich_tuan_id)
        errors = list(kiem_tra["errors"] if kiem_tra else [])
        if canh_bao_sai_cn:
            errors.extend(canh_bao_sai_cn)
        errors = list(dict.fromkeys(errors))
        return {
            "ok": True,
            "errors": errors,
            "warnings": kiem_tra["warnings"] if kiem_tra else [],
        }
    except Exception as exc:
        return {"ok": False, "message": f"Lỗi lưu kéo thả: {exc}"}


@ung_dung.post("/lich-nhap/luu")
async def luu_lich_nhap(request: Request, phien: Session = Depends(lay_phien_lam_viec)):
    try:
        payload = await request.json()
    except Exception:
        return {"ok": False, "message": "Payload không hợp lệ."}
    lich_tuan_id = payload.get("lich_tuan_id")
    if not lich_tuan_id:
        return {"ok": False, "message": "Thiếu lich_tuan_id."}
    lich = phien.query(models.LichTuan).filter(models.LichTuan.id == int(lich_tuan_id)).first()
    if not lich:
        return {"ok": False, "message": "Không tìm thấy lịch."}
    lich.trang_thai = trang_thai_chinh_thuc_tu_nhap(lich.trang_thai)
    phien.commit()
    return {"ok": True}


@ung_dung.post("/lich-nhap/huy")
async def huy_lich_nhap(request: Request, phien: Session = Depends(lay_phien_lam_viec)):
    lich_tuan_id = None
    content_type = (request.headers.get("content-type") or "").lower()
    if "application/json" in content_type:
        try:
            payload = await request.json()
            lich_tuan_id = payload.get("lich_tuan_id")
        except Exception:
            lich_tuan_id = None
    elif "text/plain" in content_type:
        try:
            raw = (await request.body()).decode("utf-8")
            payload = json.loads(raw or "{}")
            if isinstance(payload, dict):
                lich_tuan_id = payload.get("lich_tuan_id")
        except Exception:
            lich_tuan_id = None
    if not lich_tuan_id:
        return {"ok": False, "message": "Thiếu lich_tuan_id."}
    lich = phien.query(models.LichTuan).filter(models.LichTuan.id == int(lich_tuan_id)).first()
    if not lich:
        return {"ok": True}
    if la_trang_thai_nhap(lich.trang_thai):
        phien.delete(lich)
        phien.commit()
    return {"ok": True}


@ung_dung.get("/thu-nghiem")
def thu_nghiem(request: Request, ngay_bat_dau: date | None = None, phien: Session = Depends(lay_phien_lam_viec)):
    if not ngay_bat_dau:
        return tra_template(
            "thu_nghiem.html",
            {
                "request": request,
                "ket_qua": None,
                "ngay_mac_dinh": date.today(),
            },
        )

    ngay_bat_dau = ngay_bat_dau - timedelta(days=ngay_bat_dau.weekday())
    trong_so_mac_dinh = lay_trong_so(phien)
    kich_ban = [{"ten": "Mặc định", "tuy_chon": {}}]
    kich_ban.append({"ten": "Tắt tất cả trọng số", "tuy_chon": {k: 0 for k in trong_so_mac_dinh}})
    for key in trong_so_mac_dinh:
        kich_ban.append({"ten": f"Tắt {key}", "tuy_chon": {key: 0}})

    ket_qua = []
    for kb in kich_ban:
        giai = giai_lich_tuan(phien, ngay_bat_dau, luu_db=False, trong_so_tuy_chon=kb["tuy_chon"])
        thong_ke = thong_ke_phan_cong(phien, giai.phan_cong or [])
        ket_qua.append(
            {
                "ten": kb["ten"],
                "diem": giai.diem_toi_uu,
                "tong_phan_cong": giai.tong_phan_cong,
                "thong_ke": thong_ke,
                "thanh_cong": giai.thanh_cong,
                "thong_bao": giai.thong_bao,
            }
        )

    return tra_template(
        "thu_nghiem.html",
        {
            "request": request,
            "ket_qua": ket_qua,
            "ngay_mac_dinh": ngay_bat_dau,
        },
    )


@ung_dung.post("/xep-lich")
def xep_lich(
    request: Request,
    ngay_bat_dau: date = Form(...),
    phien: Session = Depends(lay_phien_lam_viec),
):
    ket_qua, ngay_bat_dau = chay_xep_lich_tuan(phien, ngay_bat_dau)
    if not ket_qua.thanh_cong:
        context = tao_context_trang_chu(
            phien,
            ngay_bat_dau,
            loi=ket_qua.thong_bao,
            thieu_nhu_cau=ket_qua.thieu_nhu_cau,
        )
        return tra_template(
            "index.html",
            {"request": request, **context},
        )
    return RedirectResponse(url=f"/?lich_tuan_id={ket_qua.lich_tuan_id}", status_code=303)


@ung_dung.post("/tu-xep-lich")
def tu_xep_lich(
    request: Request,
    ngay_bat_dau: date = Form(...),
    phien: Session = Depends(lay_phien_lam_viec),
):
    ket_qua, ngay_bat_dau = chay_tu_xep_tuan(phien, ngay_bat_dau)
    if not ket_qua.thanh_cong:
        context = tao_context_trang_chu(
            phien,
            ngay_bat_dau,
            loi=ket_qua.thong_bao,
            thieu_nhu_cau=ket_qua.thieu_nhu_cau,
        )
        return tra_template(
            "index.html",
            {"request": request, **context},
        )
    return RedirectResponse(url=f"/?lich_tuan_id={ket_qua.lich_tuan_id}", status_code=303)


@ung_dung.get("/tai-excel")
def tai_excel(lich_tuan_id: int, phien: Session = Depends(lay_phien_lam_viec)):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    du_lieu = lay_lich_hien_thi(phien, lich_tuan_id)
    if not du_lieu:
        return RedirectResponse(url="/", status_code=303)

    wb = Workbook()
    ws = wb.active
    ws.title = "Lich"

    ngay_list = du_lieu["ngay_list"]
    ws.cell(row=1, column=1, value="Nhóm / Ca")
    for i, ngay in enumerate(ngay_list, start=2):
        ws.cell(row=1, column=i, value=f"{ten_thu(ngay)} {ngay.strftime('%d/%m/%y')}")

    row = 2
    ca_theo_nhom = danh_sach_ca_theo_nhom()
    for nhom in du_lieu["nhom_list"]:
        cell = ws.cell(
            row=row,
            column=1,
            value=f"{ten_nhom_hien_thi(nhom.ten_nhom)}\n" + "\n".join(ca_theo_nhom.get(nhom.ten_nhom, [])),
        )
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        mau = (nhom.mau_nen or "#ffffff").replace("#", "")
        fill = PatternFill(start_color=mau, end_color=mau, fill_type="solid")
        for col in range(1, len(ngay_list) + 2):
            ws.cell(row=row, column=col).fill = fill
        for i, ngay in enumerate(ngay_list, start=2):
            danh_sach = du_lieu["bang"].get(nhom.id, {}).get(ngay, [])
            ten_list = [item["ten_nv"] if isinstance(item, dict) else item for item in danh_sach]
            ws.cell(row=row, column=i, value="\n".join(ten_list)).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
        row += 1

    for col in range(1, len(ngay_list) + 2):
        ws.column_dimensions[chr(64 + col)].width = 22
    for row_idx in range(1, row):
        ws.row_dimensions[row_idx].height = 72

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    data = BytesIO()
    wb.save(data)
    data.seek(0)
    return StreamingResponse(
        data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=OUTPUT_Lich_Lam_Viec.xlsx"},
    )


@ung_dung.get("/nhan-vien")
def nhan_vien(request: Request, q: str | None = None, phien: Session = Depends(lay_phien_lam_viec)):
    truy_van = phien.query(models.NhanVien)
    if q:
        truy_van = truy_van.filter(models.NhanVien.ten_nv.ilike(f"%{q}%"))
    ds_nhan_vien = truy_van.order_by(models.NhanVien.id.desc()).all()
    ds_ca = phien.query(models.CaLam).order_by(models.CaLam.id).all()
    ds_chi_nhanh = phien.query(models.ChiNhanh).order_by(models.ChiNhanh.id).all()
    ds_trong_so = phien.query(models.TrongSoUuTien).order_by(models.TrongSoUuTien.khoa).all()
    return tra_template(
        "nhan_vien.html",
        {
            "request": request,
            "ds_nhan_vien": ds_nhan_vien,
            "ds_ca": ds_ca,
            "ds_chi_nhanh": ds_chi_nhanh,
            "ds_trong_so": ds_trong_so,
            "ten_trong_so_hien_thi": ten_trong_so_hien_thi,
        },
    )


@ung_dung.get("/quan-ly")
def quan_ly(request: Request, phien: Session = Depends(lay_phien_lam_viec)):
    ds_nhan_vien = phien.query(models.NhanVien).order_by(models.NhanVien.id.desc()).all()
    ds_ca = phien.query(models.CaLam).order_by(models.CaLam.id).all()
    ds_chi_nhanh = phien.query(models.ChiNhanh).order_by(models.ChiNhanh.id).all()
    ds_trong_so = phien.query(models.TrongSoUuTien).order_by(models.TrongSoUuTien.khoa).all()
    return tra_template(
        "quan_ly.html",
        {
            "request": request,
            "ds_nhan_vien": ds_nhan_vien,
            "ds_ca": ds_ca,
            "ds_chi_nhanh": ds_chi_nhanh,
            "ds_trong_so": ds_trong_so,
            "ten_trong_so_hien_thi": ten_trong_so_hien_thi,
            "ngay_mac_dinh": date.today(),
        },
    )


@ung_dung.post("/nhan-vien")
async def tao_nhan_vien(request: Request, phien: Session = Depends(lay_phien_lam_viec)):
    form = await request.form()
    ma_nv = form.get("ma_nv") or tao_ma_nv(phien)
    danh_sach_trong_so = parse_trong_so_form(form)
    crud.tao_nhan_vien(
        phien,
        ma_nv=ma_nv,
        ten_nv=form.get("ten_nv"),
        cap_do=form.get("cap_do") or None,
        muc_uu_tien=int(form.get("muc_uu_tien") or 0),
        gio_toi_da_tuan=int(form.get("gio_toi_da_tuan") or 44),
        ghi_chu=form.get("ghi_chu") or None,
        danh_sach_vai_tro=[v.strip() for v in (form.get("vai_tro") or "").split(",") if v.strip()],
        danh_sach_ca_ua_thich=[int(v) for v in form.getlist("ca_ua_thich")],
        danh_sach_ca_tranh=[int(v) for v in form.getlist("ca_tranh")],
        danh_sach_chi_nhanh=[int(v) for v in form.getlist("chi_nhanh")],
        danh_sach_trong_so=danh_sach_trong_so,
    )
    return RedirectResponse(url="/nhan-vien", status_code=303)


@ung_dung.post("/nhan-vien/{nhan_vien_id}/cap-nhat")
async def cap_nhat_nhan_vien(nhan_vien_id: int, request: Request, phien: Session = Depends(lay_phien_lam_viec)):
    nv = phien.query(models.NhanVien).filter(models.NhanVien.id == nhan_vien_id).first()
    if not nv:
        return RedirectResponse(url="/quan-ly", status_code=303)
    form = await request.form()
    danh_sach_trong_so = parse_trong_so_form(form)
    crud.cap_nhat_nhan_vien(
        phien,
        nhan_vien=nv,
        ma_nv=form.get("ma_nv"),
        ten_nv=form.get("ten_nv"),
        cap_do=form.get("cap_do") or None,
        muc_uu_tien=int(form.get("muc_uu_tien") or nv.muc_uu_tien),
        gio_toi_da_tuan=int(form.get("gio_toi_da_tuan") or nv.gio_toi_da_tuan),
        ghi_chu=form.get("ghi_chu") or None,
        danh_sach_vai_tro=[v.strip() for v in (form.get("vai_tro") or "").split(",") if v.strip()],
        danh_sach_ca_ua_thich=[int(v) for v in form.getlist("ca_ua_thich")],
        danh_sach_ca_tranh=[int(v) for v in form.getlist("ca_tranh")],
        danh_sach_chi_nhanh=[int(v) for v in form.getlist("chi_nhanh")],
        danh_sach_trong_so=danh_sach_trong_so,
    )
    return RedirectResponse(url="/quan-ly", status_code=303)


@ung_dung.post("/nhan-vien/{nhan_vien_id}/xoa")
def xoa_nhan_vien(nhan_vien_id: int, phien: Session = Depends(lay_phien_lam_viec)):
    nv = phien.query(models.NhanVien).filter(models.NhanVien.id == nhan_vien_id).first()
    if nv:
        phien.delete(nv)
        phien.commit()
    return RedirectResponse(url="/nhan-vien", status_code=303)


@ung_dung.get("/ngay-nghi")
def ngay_nghi(request: Request, phien: Session = Depends(lay_phien_lam_viec)):
    ds_ngay_nghi = phien.query(models.NgayNghi).order_by(models.NgayNghi.ngay.desc()).all()
    ds_nhan_vien = phien.query(models.NhanVien).order_by(models.NhanVien.ten_nv).all()
    return tra_template(
        "ngay_nghi.html",
        {
            "request": request,
            "ds_ngay_nghi": ds_ngay_nghi,
            "ds_nhan_vien": ds_nhan_vien,
        },
    )


@ung_dung.post("/ngay-nghi")
async def tao_ngay_nghi(request: Request, phien: Session = Depends(lay_phien_lam_viec)):
    form = await request.form()
    try:
        nhan_vien_id = _parse_int_required(form.get("nhan_vien_id"), "nhan_vien_id")
        ngay = _parse_date_required(form.get("ngay"), "ngay")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    crud.tao_ngay_nghi(
        phien,
        nhan_vien_id=nhan_vien_id,
        ngay=ngay,
        trang_thai=form.get("trang_thai") or "OFF",
        ghi_chu=form.get("ghi_chu") or None,
    )
    return RedirectResponse(url="/ngay-nghi", status_code=303)


@ung_dung.post("/ngay-nghi/nhieu")
async def tao_ngay_nghi_nhieu(request: Request, phien: Session = Depends(lay_phien_lam_viec)):
    form = await request.form()
    nhan_vien_id = form.get("nhan_vien_id")
    ngay_list = form.getlist("ngay_list")
    if not nhan_vien_id:
        return RedirectResponse(url="/", status_code=303)
    ngay_hop_le = []
    for raw in ngay_list:
        if not raw:
            continue
        try:
            ngay_hop_le.append(datetime.fromisoformat(raw).date())
        except ValueError:
            ngay_mac_dinh = lay_thu_hai_tiep_theo()
            context = tao_context_trang_chu(
                phien,
                ngay_mac_dinh,
                loi_ngay_nghi=f"Ngày nghỉ không hợp lệ: {raw}.",
            )
            return tra_template(
                "index.html",
                {"request": request, **context},
            )
    if not ngay_hop_le:
        ngay_mac_dinh = lay_thu_hai_tiep_theo()
        context = tao_context_trang_chu(
            phien,
            ngay_mac_dinh,
            loi_ngay_nghi="Chưa chọn ngày nghỉ.",
        )
        return tra_template(
            "index.html",
            {"request": request, **context},
        )

    ngay_set = sorted(set(ngay_hop_le))
    try:
        nv_id = _parse_int_required(nhan_vien_id, "nhan_vien_id")
    except ValueError:
        return RedirectResponse(url="/", status_code=303)
    for ngay in ngay_set:
        ton_tai = (
            phien.query(models.NgayNghi)
            .filter(models.NgayNghi.nhan_vien_id == nv_id)
            .filter(models.NgayNghi.ngay == ngay)
            .first()
        )
        if ton_tai:
            continue
        phien.add(models.NgayNghi(nhan_vien_id=nv_id, ngay=ngay, trang_thai="OFF", nguon="user", ghi_chu=None))
    phien.commit()
    tuan = ngay_set[0] - timedelta(days=ngay_set[0].weekday())
    return RedirectResponse(url=f"/?ngay_bat_dau={tuan.isoformat()}", status_code=303)


@ung_dung.post("/ngay-nghi/{ngay_nghi_id}/xoa")
def xoa_ngay_nghi(
    ngay_nghi_id: int,
    redirect_to: str | None = Form(None),
    phien: Session = Depends(lay_phien_lam_viec),
):
    nn = phien.query(models.NgayNghi).filter(models.NgayNghi.id == ngay_nghi_id).first()
    if nn:
        phien.delete(nn)
        phien.commit()
    return RedirectResponse(url=redirect_to or "/ngay-nghi", status_code=303)


@ung_dung.post("/ngay-nghi/xoa-tat-ca")
def xoa_ngay_nghi_tat_ca(
    request: Request,
    redirect_to: str | None = Form(None),
    admin_token: str | None = Form(None),
    phien: Session = Depends(lay_phien_lam_viec),
):
    kiem_tra_admin_token(request, admin_token)
    phien.query(models.NgayNghi).delete()
    phien.commit()
    return RedirectResponse(url=redirect_to or "/ngay-nghi", status_code=303)


@ung_dung.get("/nhu-cau-ca")
def nhu_cau_ca(request: Request, phien: Session = Depends(lay_phien_lam_viec)):
    ds_nhu_cau = phien.query(models.NhuCauCa).order_by(models.NhuCauCa.ngay.desc()).all()
    ds_chi_nhanh = phien.query(models.ChiNhanh).order_by(models.ChiNhanh.id).all()
    ds_ca = phien.query(models.CaLam).order_by(models.CaLam.id).all()
    ds_vai_tro = phien.query(models.VaiTro).order_by(models.VaiTro.id).all()
    return tra_template(
        "nhu_cau_ca.html",
        {
            "request": request,
            "ds_nhu_cau": ds_nhu_cau,
            "ds_chi_nhanh": ds_chi_nhanh,
            "ds_ca": ds_ca,
            "ds_vai_tro": ds_vai_tro,
        },
    )


@ung_dung.post("/nhu-cau-ca")
async def tao_nhu_cau_ca(request: Request, phien: Session = Depends(lay_phien_lam_viec)):
    form = await request.form()
    try:
        chi_nhanh_id = _parse_int_or_none(form.get("chi_nhanh_id"), "chi_nhanh_id")
        vai_tro_id = _parse_int_or_none(form.get("vai_tro_yeu_cau_id"), "vai_tro_yeu_cau_id")
        ca_id = _parse_int_required(form.get("ca_id"), "ca_id")
        so_nguoi_can = _parse_int_required(form.get("so_nguoi_can"), "so_nguoi_can")
        do_quan_trong = _parse_int_or_none(form.get("do_quan_trong"), "do_quan_trong")
        senior_toi_thieu = _parse_int_or_none(form.get("senior_toi_thieu"), "senior_toi_thieu")
        ngay = _parse_date_required(form.get("ngay"), "ngay")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    crud.tao_nhu_cau_ca(
        phien,
        ngay=ngay,
        chi_nhanh_id=chi_nhanh_id,
        ca_id=ca_id,
        so_nguoi_can=so_nguoi_can,
        vai_tro_yeu_cau_id=vai_tro_id,
        do_quan_trong=do_quan_trong,
        senior_toi_thieu=senior_toi_thieu,
    )
    return RedirectResponse(url="/nhu-cau-ca", status_code=303)


@ung_dung.post("/nhu-cau-ca/{nhu_cau_id}/xoa")
def xoa_nhu_cau_ca(nhu_cau_id: int, phien: Session = Depends(lay_phien_lam_viec)):
    nc = phien.query(models.NhuCauCa).filter(models.NhuCauCa.id == nhu_cau_id).first()
    if nc:
        phien.delete(nc)
        phien.commit()
    return RedirectResponse(url="/nhu-cau-ca", status_code=303)


@ung_dung.get("/trong-so")
def trong_so(request: Request, phien: Session = Depends(lay_phien_lam_viec)):
    ds_trong_so = phien.query(models.TrongSoUuTien).order_by(models.TrongSoUuTien.khoa).all()
    return tra_template(
        "trong_so.html",
        {
            "request": request,
            "ds_trong_so": ds_trong_so,
            "ten_trong_so_hien_thi": ten_trong_so_hien_thi,
        },
    )


@ung_dung.post("/trong-so")
async def cap_nhat_trong_so(request: Request, phien: Session = Depends(lay_phien_lam_viec)):
    form = await request.form()
    try:
        gia_tri = _parse_int_required(form.get("gia_tri"), "gia_tri")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    crud.tao_hoac_cap_nhat_trong_so(
        phien,
        khoa=form.get("khoa"),
        gia_tri=gia_tri,
    )
    return RedirectResponse(url="/trong-so", status_code=303)


@ung_dung.post("/trong-so/{trong_so_id}/xoa")
def xoa_trong_so(trong_so_id: int, phien: Session = Depends(lay_phien_lam_viec)):
    ts = phien.query(models.TrongSoUuTien).filter(models.TrongSoUuTien.id == trong_so_id).first()
    if ts:
        phien.delete(ts)
        phien.commit()
    return RedirectResponse(url="/trong-so", status_code=303)
