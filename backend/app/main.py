# -*- coding: utf-8 -*-
from __future__ import annotations

from datetime import date, datetime, timedelta
import time
from io import BytesIO

from fastapi import Depends, FastAPI, Form, Request
from fastapi.responses import RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy import func
from sqlalchemy.orm import Session

from backend.app import crud, models
from backend.app.db import CoSo, dong_co, lay_phien_lam_viec
from backend.app.scheduler.solver import (
    danh_sach_ngay_trong_tuan,
    giai_lich_tuan,
    lay_trong_so,
    tao_nhu_cau_chich_ngoai,
)
from backend.app.seed import tao_du_lieu_mau


ung_dung = FastAPI(title="Lich lam viec")
ung_dung.mount("/static", StaticFiles(directory="backend/app/static"), name="static")
giao_dien = Jinja2Templates(directory="backend/app/templates")


@ung_dung.on_event("startup")
def khoi_tao():
    for _ in range(10):
        try:
            CoSo.metadata.create_all(bind=dong_co)
            with next(lay_phien_lam_viec()) as phien:
                tao_du_lieu_mau(phien)
                cap_nhat_nhom_off(phien)
            return
        except Exception:
            time.sleep(1)
    CoSo.metadata.create_all(bind=dong_co)
    with next(lay_phien_lam_viec()) as phien:
        tao_du_lieu_mau(phien)
        cap_nhat_nhom_off(phien)


def ten_thu(ngay: date) -> str:
    thu = ["Hai", "Ba", "Tư", "Năm", "Sáu", "Bảy", "Chủ nhật"]
    return thu[ngay.weekday()]


def ten_nhom_hien_thi(ten_nhom: str) -> str:
    if ten_nhom == "CN":
        return "Chích ngoài"
    if ten_nhom in {"Spa", "OFF"}:
        return "OFF (Ngày nghỉ)"
    return ten_nhom


def ten_trong_so_hien_thi(khoa: str) -> str:
    mapping = {
        "uu_tien_ca_ua_thich": "Ưu tiên ca ưa thích",
        "phat_ca_tranh": "Phạt ca tránh",
        "cong_bang_chich_ngoai": "Công bằng chích ngoài",
        "cong_bang_cuoi_tuan": "Công bằng cuối tuần",
        "uu_tien_ca_quan_trong": "Ưu tiên ca quan trọng",
        "han_che_ca_muon_sang": "Hạn chế ca muộn/sáng",
        "bat_buoc_chich_ngoai": "Bắt buộc chích ngoài",
    }
    return mapping.get(khoa, khoa)


def dinh_dang_ngay(ngay: date) -> str:
    return ngay.strftime("%d/%m/%Y")


def tao_ma_nv(phien: Session) -> str:
    max_id = phien.query(func.max(models.NhanVien.id)).scalar() or 0
    return f"BS{max_id + 1:02d}"


def parse_trong_so_form(form) -> list[int]:
    ket_qua = []
    for idx in range(1, 4):
        ts_id = form.get(f"trong_so_{idx}_id")
        if not ts_id:
            continue
        ket_qua.append(int(ts_id))
    return ket_qua


def danh_sach_ca_theo_nhom() -> dict[str, list[str]]:
    return {
        "326TTV": ["8h-19h", "8h30-19h30", "9h-20h", "10h-21h"],
        "197LT5": ["8h-19h", "9h-20h", "10h-21h"],
        "796ADV": ["9h-20h"],
        "CN": ["9h-20h"],
        "OFF": ["Nghỉ"],
    }


def la_nhom_off(ten_nhom: str | None) -> bool:
    if not ten_nhom:
        return False
    return ten_nhom.strip().lower() in {"spa", "off", "nghi", "ngay nghi", "off (ngay nghi)"}


def cap_nhat_nhom_off(phien: Session):
    nhom_off = phien.query(models.NhomHienThi).filter(models.NhomHienThi.ten_nhom == "OFF").first()
    nhom_spa = phien.query(models.NhomHienThi).filter(models.NhomHienThi.ten_nhom == "Spa").first()
    if nhom_spa and not nhom_off:
        nhom_spa.ten_nhom = "OFF"
        phien.flush()
        phien.query(models.MappingNhom).filter(models.MappingNhom.nhom_hien_thi_id == nhom_spa.id).delete()
        phien.commit()
        return
    if nhom_spa and nhom_off:
        phien.query(models.MappingNhom).filter(models.MappingNhom.nhom_hien_thi_id == nhom_spa.id).delete()
        phien.delete(nhom_spa)
        phien.commit()
        return
    if nhom_off:
        phien.query(models.MappingNhom).filter(models.MappingNhom.nhom_hien_thi_id == nhom_off.id).delete()
        phien.commit()


def lay_lich_hien_thi(phien: Session, lich_tuan_id: int | None):
    lich_tuan = None
    if lich_tuan_id:
        lich_tuan = phien.query(models.LichTuan).filter(models.LichTuan.id == lich_tuan_id).first()
    if not lich_tuan:
        lich_tuan = phien.query(models.LichTuan).order_by(models.LichTuan.id.desc()).first()
    if not lich_tuan:
        return None

    nhom_list = phien.query(models.NhomHienThi).order_by(models.NhomHienThi.id).all()
    lich_ct = (
        phien.query(models.LichChiTiet)
        .filter(models.LichChiTiet.lich_tuan_id == lich_tuan.id)
        .all()
    )
    bang = {
        nhom.id: {ngay: [] for ngay in danh_sach_ngay_trong_tuan(lich_tuan.ngay_bat_dau)}
        for nhom in nhom_list
    }
    for ct in lich_ct:
        if ct.nhom_hien_thi_id not in bang:
            continue
        bang[ct.nhom_hien_thi_id][ct.ngay].append(
            {
                "id": ct.id,
                "ten_nv": ct.nhan_vien.ten_nv,
                "nhan_vien_id": ct.nhan_vien_id,
            }
        )

    return {
        "lich_tuan": lich_tuan,
        "nhom_list": nhom_list,
        "ngay_list": danh_sach_ngay_trong_tuan(lich_tuan.ngay_bat_dau),
        "bang": bang,
    }


def kiem_tra_lich(phien: Session, lich_tuan_id: int | None):
    lich_tuan = None
    if lich_tuan_id:
        lich_tuan = phien.query(models.LichTuan).filter(models.LichTuan.id == lich_tuan_id).first()
    if not lich_tuan:
        lich_tuan = phien.query(models.LichTuan).order_by(models.LichTuan.id.desc()).first()
    if not lich_tuan:
        return None

    ngay_list = danh_sach_ngay_trong_tuan(lich_tuan.ngay_bat_dau)
    nhu_cau = (
        phien.query(models.NhuCauCa)
        .filter(models.NhuCauCa.ngay >= ngay_list[0])
        .filter(models.NhuCauCa.ngay <= ngay_list[-1])
        .all()
    )
    trong_so = lay_trong_so(phien)
    if trong_so.get("bat_buoc_chich_ngoai", 0) > 0:
        nhu_cau = list(nhu_cau) + tao_nhu_cau_chich_ngoai(phien, ngay_list)
    lich_ct = (
        phien.query(models.LichChiTiet)
        .filter(models.LichChiTiet.lich_tuan_id == lich_tuan.id)
        .all()
    )

    dem_lich = {}
    for ct in lich_ct:
        key = (ct.ngay, ct.ca_id, ct.chi_nhanh_id)
        dem_lich[key] = dem_lich.get(key, 0) + 1

    errors = []
    warnings = []
    for nc in nhu_cau:
        key = (nc.ngay, nc.ca_id, nc.chi_nhanh_id)
        da_xep = dem_lich.get(key, 0)
        if da_xep < nc.so_nguoi_can:
            ten_chi_nhanh = nc.chi_nhanh.ten_chi_nhanh if nc.chi_nhanh else "Chích ngoài"
            ten_ca = nc.ca_lam.ten_ca if nc.ca_lam else ""
            errors.append(
                f"Thiếu {nc.so_nguoi_can - da_xep} người: {ten_chi_nhanh} {dinh_dang_ngay(nc.ngay)} ca {ten_ca}"
            )

    ca_map = {ca.id: ca for ca in phien.query(models.CaLam).all()}
    gio_nv = {}
    ca_trong_ngay = {}
    for ct in lich_ct:
        ca = ca_map.get(ct.ca_id)
        so_gio = ca.so_gio if ca else 0
        gio_nv[ct.nhan_vien_id] = gio_nv.get(ct.nhan_vien_id, 0) + so_gio
        ngay_key = (ct.nhan_vien_id, ct.ngay)
        ca_trong_ngay[ngay_key] = ca_trong_ngay.get(ngay_key, 0) + 1

    nhan_vien_map = {nv.id: nv for nv in phien.query(models.NhanVien).all()}
    for nv_id, tong_gio in gio_nv.items():
        nv = nhan_vien_map.get(nv_id)
        if nv and tong_gio > nv.gio_toi_da_tuan:
            warnings.append(f"Vượt giờ tuần: {nv.ten_nv} ({tong_gio}/{nv.gio_toi_da_tuan} giờ)")

    for (nv_id, ngay), so_ca in ca_trong_ngay.items():
        if so_ca > 1:
            nv = nhan_vien_map.get(nv_id)
            ten = nv.ten_nv if nv else f"NV {nv_id}"
            errors.append(f"Trùng ca trong ngày: {ten} vào {dinh_dang_ngay(ngay)}")

    return {
        "lich_tuan": lich_tuan,
        "errors": errors,
        "warnings": warnings,
    }


def thong_ke_phan_cong(phien: Session, phan_cong: list[dict]) -> dict[str, int]:
    if not phan_cong:
        return {
            "tong_ca": 0,
            "ca_ua_thich": 0,
            "ca_tranh": 0,
            "chich_ngoai": 0,
            "cuoi_tuan": 0,
            "ca_quan_trong": 0,
        }
    nv_map = {nv.id: nv for nv in phien.query(models.NhanVien).all()}
    dem = {
        "tong_ca": 0,
        "ca_ua_thich": 0,
        "ca_tranh": 0,
        "chich_ngoai": 0,
        "cuoi_tuan": 0,
        "ca_quan_trong": 0,
    }
    for pc in phan_cong:
        dem["tong_ca"] += 1
        nv = nv_map.get(pc["nhan_vien_id"])
        if nv:
            if pc["ca_id"] in {ca.id for ca in nv.ca_ua_thich}:
                dem["ca_ua_thich"] += 1
            if pc["ca_id"] in {ca.id for ca in nv.ca_tranh}:
                dem["ca_tranh"] += 1
        if pc.get("la_chich_ngoai"):
            dem["chich_ngoai"] += 1
        if pc["ngay"].weekday() >= 5:
            dem["cuoi_tuan"] += 1
        if pc.get("do_quan_trong", 0) > 0:
            dem["ca_quan_trong"] += 1
    return dem


@ung_dung.get("/")
def trang_chu(
    request: Request,
    lich_tuan_id: int | None = None,
    sap_xep: str | None = None,
    phien: Session = Depends(lay_phien_lam_viec),
):
    du_lieu_lich = lay_lich_hien_thi(phien, lich_tuan_id)
    ngay_mac_dinh = du_lieu_lich["lich_tuan"].ngay_bat_dau if du_lieu_lich else date.today()
    ds_nhan_vien = phien.query(models.NhanVien).order_by(models.NhanVien.ten_nv).all()
    ngay_bat_dau = ngay_mac_dinh - timedelta(days=ngay_mac_dinh.weekday())
    ngay_list_nghi = danh_sach_ngay_trong_tuan(ngay_bat_dau)
    ds_ngay_nghi = (
        phien.query(models.NgayNghi)
        .filter(models.NgayNghi.ngay >= ngay_list_nghi[0])
        .filter(models.NgayNghi.ngay <= ngay_list_nghi[-1])
        .order_by(models.NgayNghi.ngay.asc())
        .all()
    )
    ngay_nghi_theo_ngay = {ngay: [] for ngay in ngay_list_nghi}
    for nn in ds_ngay_nghi:
        if nn.ngay in ngay_nghi_theo_ngay:
            ngay_nghi_theo_ngay[nn.ngay].append(nn)
    return giao_dien.TemplateResponse(
        "index.html",
        {
            "request": request,
            "lich": du_lieu_lich,
            "danh_sach_ca_nhom": danh_sach_ca_theo_nhom(),
            "ten_thu": ten_thu,
            "ten_nhom_hien_thi": ten_nhom_hien_thi,
            "ngay_mac_dinh": ngay_mac_dinh,
            "ds_nhan_vien": ds_nhan_vien,
            "ds_ngay_nghi": ds_ngay_nghi,
            "ngay_list_nghi": ngay_list_nghi,
            "ngay_nghi_theo_ngay": ngay_nghi_theo_ngay,
        },
    )


@ung_dung.get("/kiem-tra")
def kiem_tra(request: Request, lich_tuan_id: int | None = None, phien: Session = Depends(lay_phien_lam_viec)):
    ket_qua = kiem_tra_lich(phien, lich_tuan_id)
    if not ket_qua:
        return RedirectResponse(url="/", status_code=303)
    return giao_dien.TemplateResponse(
        "kiem_tra.html",
        {
            "request": request,
            "ket_qua": ket_qua,
        },
    )


@ung_dung.post("/cap-nhat-lich")
async def cap_nhat_lich(request: Request, phien: Session = Depends(lay_phien_lam_viec)):
    try:
        payload = await request.json()
        lich_tuan_id = payload.get("lich_tuan_id")
        changes = payload.get("changes", [])
        if not lich_tuan_id or not changes:
            return {"ok": False, "message": "Không có thay đổi."}

        lich_ct = (
            phien.query(models.LichChiTiet)
            .filter(models.LichChiTiet.lich_tuan_id == lich_tuan_id)
            .all()
        )
        nhom_list = phien.query(models.NhomHienThi).all()
        nhom_map = {nhom.id: nhom for nhom in nhom_list}
        off_nhom_ids = {nhom.id for nhom in nhom_list if la_nhom_off(nhom.ten_nhom)}
        mapping_list = phien.query(models.MappingNhom).all()
        mapping_map = {(m.nhom_hien_thi_id, m.ca_id): m.chi_nhanh_id for m in mapping_list}
        lich_map = {ct.id: ct for ct in lich_ct}
        du_kien = {}
        for ct in lich_ct:
            du_kien[ct.id] = {
                "ngay": ct.ngay,
                "nhom_hien_thi_id": ct.nhom_hien_thi_id,
                "nhan_vien_id": ct.nhan_vien_id,
                "chi_nhanh_id": ct.chi_nhanh_id,
                "ca_id": ct.ca_id,
            }

        for item in changes:
            lich_id = item.get("id")
            ngay = item.get("ngay")
            nhom_id = item.get("nhom_hien_thi_id")
            if not lich_id or not ngay or not nhom_id or lich_id not in du_kien:
                continue
            du_kien[lich_id]["ngay"] = datetime.fromisoformat(ngay).date()
            nhom_id = int(nhom_id)
            du_kien[lich_id]["nhom_hien_thi_id"] = nhom_id
            if nhom_id in off_nhom_ids:
                du_kien[lich_id]["chi_nhanh_id"] = None
                du_kien[lich_id]["ca_id"] = None
                continue
            ca_id = du_kien[lich_id]["ca_id"]
            if ca_id:
                chi_nhanh_id = mapping_map.get((nhom_id, ca_id))
                if (nhom_id, ca_id) not in mapping_map:
                    nhom_obj = nhom_map.get(nhom_id)
                    nhom_ten = nhom_obj.ten_nhom if nhom_obj else ""
                    return {
                        "ok": False,
                        "message": f"Không có mapping chi nhánh cho nhóm {nhom_ten}.",
                    }
                du_kien[lich_id]["chi_nhanh_id"] = chi_nhanh_id

        dem_trung = {}
        for info in du_kien.values():
            key = (info["nhan_vien_id"], info["ngay"])
            dem_trung[key] = dem_trung.get(key, 0) + 1
        bi_trung = [key for key, so_ca in dem_trung.items() if so_ca > 1]
        if bi_trung:
            nv_map = {nv.id: nv for nv in phien.query(models.NhanVien).all()}
            chi_nhanh_map = {cn.id: cn for cn in phien.query(models.ChiNhanh).all()}
            ca_map = {ca.id: ca for ca in phien.query(models.CaLam).all()}
            danh_sach = []
            for nv_id, ngay in bi_trung:
                nv = nv_map.get(nv_id)
                ten = nv.ten_nv if nv else nv_id
                chi_tiet = set()
                for info in du_kien.values():
                    if info["nhan_vien_id"] == nv_id and info["ngay"] == ngay:
                        cn = chi_nhanh_map.get(info["chi_nhanh_id"])
                        ca = ca_map.get(info["ca_id"])
                        ten_cn = cn.ten_chi_nhanh if cn else "Chích ngoài"
                        ten_ca = ca.ten_ca if ca else ""
                        chi_tiet.add(f"{ten_cn} {ten_ca}".strip())
                danh_sach.append(f"{ten} {dinh_dang_ngay(ngay)} ({'; '.join(sorted(chi_tiet))})")
            danh_sach = ", ".join(danh_sach)
            return {"ok": False, "message": f"Trùng ca trong ngày: {danh_sach}."}

        nv_map = {nv.id: nv for nv in phien.query(models.NhanVien).all()}
        chi_nhanh_map = {cn.id: cn for cn in phien.query(models.ChiNhanh).all()}
        for info in du_kien.values():
            nv = nv_map.get(info["nhan_vien_id"])
            if not nv or not nv.chi_nhanh or not info["chi_nhanh_id"]:
                continue
            if info["chi_nhanh_id"] not in {cn.id for cn in nv.chi_nhanh}:
                cn = chi_nhanh_map.get(info["chi_nhanh_id"])
                ten_cn = cn.ten_chi_nhanh if cn else ""
                return {
                    "ok": False,
                    "message": f"Nhân viên {nv.ten_nv} không thuộc chi nhánh {ten_cn}.",
                }

        off_sau = {
            (info["nhan_vien_id"], info["ngay"])
            for info in du_kien.values()
            if info["nhom_hien_thi_id"] in off_nhom_ids
        }

        for lich_id, info in du_kien.items():
            ct = lich_map.get(lich_id)
            if not ct:
                continue
            ct.ngay = info["ngay"]
            ct.nhom_hien_thi_id = info["nhom_hien_thi_id"]
            ct.chi_nhanh_id = info["chi_nhanh_id"]
            ct.ca_id = info["ca_id"]

        ngay_set = {info["ngay"] for info in du_kien.values()}
        ngay_nghi_hien_tai = (
            phien.query(models.NgayNghi)
            .filter(models.NgayNghi.ngay.in_(ngay_set))
            .all()
        )
        ngay_nghi_map = {(nn.nhan_vien_id, nn.ngay): nn for nn in ngay_nghi_hien_tai}
        for nv_id, ngay in off_sau:
            if (nv_id, ngay) not in ngay_nghi_map:
                phien.add(
                    models.NgayNghi(
                        nhan_vien_id=nv_id,
                        ngay=ngay,
                        trang_thai="OFF",
                        ghi_chu=None,
                    )
                )
        for key, nn in ngay_nghi_map.items():
            if key not in off_sau:
                phien.delete(nn)
        phien.commit()
        kiem_tra = kiem_tra_lich(phien, lich_tuan_id)
        return {
            "ok": True,
            "errors": kiem_tra["errors"] if kiem_tra else [],
            "warnings": kiem_tra["warnings"] if kiem_tra else [],
        }
    except Exception as exc:
        return {"ok": False, "message": f"Lỗi lưu kéo thả: {exc}"}


@ung_dung.get("/thu-nghiem")
def thu_nghiem(request: Request, ngay_bat_dau: date | None = None, phien: Session = Depends(lay_phien_lam_viec)):
    if not ngay_bat_dau:
        return giao_dien.TemplateResponse(
            "thu_nghiem.html",
            {
                "request": request,
                "ket_qua": None,
                "ngay_mac_dinh": date.today(),
            },
        )

    ngay_bat_dau = ngay_bat_dau - timedelta(days=ngay_bat_dau.weekday())
    trong_so_mac_dinh = lay_trong_so(phien)
    kich_ban = [{"ten": "Mặc định", "tuy_chon": {}}]
    kich_ban.append({"ten": "Tắt tất cả trọng số", "tuy_chon": {k: 0 for k in trong_so_mac_dinh}})
    for key in trong_so_mac_dinh:
        kich_ban.append({"ten": f"Tắt {key}", "tuy_chon": {key: 0}})

    ket_qua = []
    for kb in kich_ban:
        giai = giai_lich_tuan(phien, ngay_bat_dau, luu_db=False, trong_so_tuy_chon=kb["tuy_chon"])
        thong_ke = thong_ke_phan_cong(phien, giai.phan_cong or [])
        ket_qua.append(
            {
                "ten": kb["ten"],
                "diem": giai.diem_toi_uu,
                "tong_phan_cong": giai.tong_phan_cong,
                "thong_ke": thong_ke,
                "thanh_cong": giai.thanh_cong,
                "thong_bao": giai.thong_bao,
            }
        )

    return giao_dien.TemplateResponse(
        "thu_nghiem.html",
        {
            "request": request,
            "ket_qua": ket_qua,
            "ngay_mac_dinh": ngay_bat_dau,
        },
    )


@ung_dung.post("/xep-lich")
def xep_lich(
    request: Request,
    ngay_bat_dau: date = Form(...),
    phien: Session = Depends(lay_phien_lam_viec),
):
    ngay_bat_dau = ngay_bat_dau - timedelta(days=ngay_bat_dau.weekday())
    ket_qua = giai_lich_tuan(phien, ngay_bat_dau)
    if not ket_qua.thanh_cong:
        return giao_dien.TemplateResponse(
            "index.html",
            {
                "request": request,
                "loi": ket_qua.thong_bao,
                "thieu_nhu_cau": ket_qua.thieu_nhu_cau,
                "lich": lay_lich_hien_thi(phien, None),
                "danh_sach_ca_nhom": danh_sach_ca_theo_nhom(),
                "ten_thu": ten_thu,
                "ten_nhom_hien_thi": ten_nhom_hien_thi,
                "ngay_mac_dinh": ngay_bat_dau,
                "ds_nhan_vien": phien.query(models.NhanVien).order_by(models.NhanVien.ten_nv).all(),
                "ds_ngay_nghi": phien.query(models.NgayNghi).order_by(models.NgayNghi.ngay.desc()).all(),
                "ngay_list_nghi": danh_sach_ngay_trong_tuan(ngay_bat_dau),
                "ngay_nghi_theo_ngay": {},
            },
        )
    return RedirectResponse(url=f"/?lich_tuan_id={ket_qua.lich_tuan_id}", status_code=303)


@ung_dung.get("/tai-excel")
def tai_excel(lich_tuan_id: int, phien: Session = Depends(lay_phien_lam_viec)):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    du_lieu = lay_lich_hien_thi(phien, lich_tuan_id)
    if not du_lieu:
        return RedirectResponse(url="/", status_code=303)

    wb = Workbook()
    ws = wb.active
    ws.title = "Lich"

    ngay_list = du_lieu["ngay_list"]
    ws.cell(row=1, column=1, value="Nhóm / Ca")
    for i, ngay in enumerate(ngay_list, start=2):
        ws.cell(row=1, column=i, value=f"{ten_thu(ngay)} {ngay.strftime('%d/%m/%y')}")

    row = 2
    ca_theo_nhom = danh_sach_ca_theo_nhom()
    for nhom in du_lieu["nhom_list"]:
        cell = ws.cell(
            row=row,
            column=1,
            value=f"{ten_nhom_hien_thi(nhom.ten_nhom)}\n" + "\n".join(ca_theo_nhom.get(nhom.ten_nhom, [])),
        )
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        mau = (nhom.mau_nen or "#ffffff").replace("#", "")
        fill = PatternFill(start_color=mau, end_color=mau, fill_type="solid")
        for col in range(1, len(ngay_list) + 2):
            ws.cell(row=row, column=col).fill = fill
        for i, ngay in enumerate(ngay_list, start=2):
            danh_sach = du_lieu["bang"].get(nhom.id, {}).get(ngay, [])
            ten_list = [item["ten_nv"] if isinstance(item, dict) else item for item in danh_sach]
            ws.cell(row=row, column=i, value="\n".join(ten_list)).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
        row += 1

    for col in range(1, len(ngay_list) + 2):
        ws.column_dimensions[chr(64 + col)].width = 22
    for row_idx in range(1, row):
        ws.row_dimensions[row_idx].height = 72

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    data = BytesIO()
    wb.save(data)
    data.seek(0)
    return StreamingResponse(
        data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=OUTPUT_Lich_Lam_Viec.xlsx"},
    )


@ung_dung.get("/nhan-vien")
def nhan_vien(request: Request, q: str | None = None, phien: Session = Depends(lay_phien_lam_viec)):
    truy_van = phien.query(models.NhanVien)
    if q:
        truy_van = truy_van.filter(models.NhanVien.ten_nv.ilike(f"%{q}%"))
    ds_nhan_vien = truy_van.order_by(models.NhanVien.id.desc()).all()
    ds_ca = phien.query(models.CaLam).order_by(models.CaLam.id).all()
    ds_chi_nhanh = phien.query(models.ChiNhanh).order_by(models.ChiNhanh.id).all()
    ds_trong_so = phien.query(models.TrongSoUuTien).order_by(models.TrongSoUuTien.khoa).all()
    return giao_dien.TemplateResponse(
        "nhan_vien.html",
        {
            "request": request,
            "ds_nhan_vien": ds_nhan_vien,
            "ds_ca": ds_ca,
            "ds_chi_nhanh": ds_chi_nhanh,
            "ds_trong_so": ds_trong_so,
            "ten_trong_so_hien_thi": ten_trong_so_hien_thi,
        },
    )


@ung_dung.get("/quan-ly")
def quan_ly(request: Request, phien: Session = Depends(lay_phien_lam_viec)):
    ds_nhan_vien = phien.query(models.NhanVien).order_by(models.NhanVien.id.desc()).all()
    ds_ca = phien.query(models.CaLam).order_by(models.CaLam.id).all()
    ds_chi_nhanh = phien.query(models.ChiNhanh).order_by(models.ChiNhanh.id).all()
    ds_trong_so = phien.query(models.TrongSoUuTien).order_by(models.TrongSoUuTien.khoa).all()
    return giao_dien.TemplateResponse(
        "quan_ly.html",
        {
            "request": request,
            "ds_nhan_vien": ds_nhan_vien,
            "ds_ca": ds_ca,
            "ds_chi_nhanh": ds_chi_nhanh,
            "ds_trong_so": ds_trong_so,
            "ten_trong_so_hien_thi": ten_trong_so_hien_thi,
            "ngay_mac_dinh": date.today(),
        },
    )


@ung_dung.post("/nhan-vien")
async def tao_nhan_vien(request: Request, phien: Session = Depends(lay_phien_lam_viec)):
    form = await request.form()
    ma_nv = form.get("ma_nv") or tao_ma_nv(phien)
    danh_sach_trong_so = parse_trong_so_form(form)
    crud.tao_nhan_vien(
        phien,
        ma_nv=ma_nv,
        ten_nv=form.get("ten_nv"),
        cap_do=form.get("cap_do") or None,
        muc_uu_tien=int(form.get("muc_uu_tien") or 0),
        gio_toi_da_tuan=int(form.get("gio_toi_da_tuan") or 44),
        ghi_chu=form.get("ghi_chu") or None,
        danh_sach_vai_tro=[v.strip() for v in (form.get("vai_tro") or "").split(",") if v.strip()],
        danh_sach_ca_ua_thich=[int(v) for v in form.getlist("ca_ua_thich")],
        danh_sach_ca_tranh=[int(v) for v in form.getlist("ca_tranh")],
        danh_sach_chi_nhanh=[int(v) for v in form.getlist("chi_nhanh")],
        danh_sach_trong_so=danh_sach_trong_so,
    )
    return RedirectResponse(url="/nhan-vien", status_code=303)


@ung_dung.post("/nhan-vien/{nhan_vien_id}/cap-nhat")
async def cap_nhat_nhan_vien(nhan_vien_id: int, request: Request, phien: Session = Depends(lay_phien_lam_viec)):
    nv = phien.query(models.NhanVien).filter(models.NhanVien.id == nhan_vien_id).first()
    if not nv:
        return RedirectResponse(url="/quan-ly", status_code=303)
    form = await request.form()
    danh_sach_trong_so = parse_trong_so_form(form)
    crud.cap_nhat_nhan_vien(
        phien,
        nhan_vien=nv,
        ma_nv=form.get("ma_nv"),
        ten_nv=form.get("ten_nv"),
        cap_do=form.get("cap_do") or None,
        muc_uu_tien=int(form.get("muc_uu_tien") or nv.muc_uu_tien),
        gio_toi_da_tuan=int(form.get("gio_toi_da_tuan") or nv.gio_toi_da_tuan),
        ghi_chu=form.get("ghi_chu") or None,
        danh_sach_vai_tro=[v.strip() for v in (form.get("vai_tro") or "").split(",") if v.strip()],
        danh_sach_ca_ua_thich=[int(v) for v in form.getlist("ca_ua_thich")],
        danh_sach_ca_tranh=[int(v) for v in form.getlist("ca_tranh")],
        danh_sach_chi_nhanh=[int(v) for v in form.getlist("chi_nhanh")],
        danh_sach_trong_so=danh_sach_trong_so,
    )
    return RedirectResponse(url="/quan-ly", status_code=303)


@ung_dung.post("/nhan-vien/{nhan_vien_id}/xoa")
def xoa_nhan_vien(nhan_vien_id: int, phien: Session = Depends(lay_phien_lam_viec)):
    nv = phien.query(models.NhanVien).filter(models.NhanVien.id == nhan_vien_id).first()
    if nv:
        phien.delete(nv)
        phien.commit()
    return RedirectResponse(url="/nhan-vien", status_code=303)


@ung_dung.get("/ngay-nghi")
def ngay_nghi(request: Request, phien: Session = Depends(lay_phien_lam_viec)):
    ds_ngay_nghi = phien.query(models.NgayNghi).order_by(models.NgayNghi.ngay.desc()).all()
    ds_nhan_vien = phien.query(models.NhanVien).order_by(models.NhanVien.ten_nv).all()
    return giao_dien.TemplateResponse(
        "ngay_nghi.html",
        {
            "request": request,
            "ds_ngay_nghi": ds_ngay_nghi,
            "ds_nhan_vien": ds_nhan_vien,
        },
    )


@ung_dung.post("/ngay-nghi")
async def tao_ngay_nghi(request: Request, phien: Session = Depends(lay_phien_lam_viec)):
    form = await request.form()
    crud.tao_ngay_nghi(
        phien,
        nhan_vien_id=int(form.get("nhan_vien_id")),
        ngay=datetime.fromisoformat(form.get("ngay")).date(),
        trang_thai=form.get("trang_thai") or "OFF",
        ghi_chu=form.get("ghi_chu") or None,
    )
    return RedirectResponse(url="/ngay-nghi", status_code=303)


@ung_dung.post("/ngay-nghi/nhieu")
async def tao_ngay_nghi_nhieu(request: Request, phien: Session = Depends(lay_phien_lam_viec)):
    form = await request.form()
    nhan_vien_id = form.get("nhan_vien_id")
    ngay_list = form.getlist("ngay_list")
    if not nhan_vien_id:
        return RedirectResponse(url="/", status_code=303)
    ngay_hop_le = []
    for raw in ngay_list:
        if not raw:
            continue
        try:
            ngay_hop_le.append(datetime.fromisoformat(raw).date())
        except ValueError:
            du_lieu_lich = lay_lich_hien_thi(phien, None)
            return giao_dien.TemplateResponse(
                "index.html",
                {
                    "request": request,
                    "lich": du_lieu_lich,
                    "danh_sach_ca_nhom": danh_sach_ca_theo_nhom(),
                    "ten_thu": ten_thu,
                    "ten_nhom_hien_thi": ten_nhom_hien_thi,
                    "ngay_mac_dinh": date.today(),
                    "ds_nhan_vien": phien.query(models.NhanVien).order_by(models.NhanVien.ten_nv).all(),
                    "ds_ngay_nghi": phien.query(models.NgayNghi).order_by(models.NgayNghi.ngay.desc()).all(),
                    "sap_xep_ngay_nghi": "ngay_moi",
                    "loi_ngay_nghi": f"Ngày nghỉ không hợp lệ: {raw}.",
                },
            )
    if not ngay_hop_le:
        du_lieu_lich = lay_lich_hien_thi(phien, None)
        return giao_dien.TemplateResponse(
            "index.html",
            {
                "request": request,
                "lich": du_lieu_lich,
                "danh_sach_ca_nhom": danh_sach_ca_theo_nhom(),
                "ten_thu": ten_thu,
                "ten_nhom_hien_thi": ten_nhom_hien_thi,
                "ngay_mac_dinh": date.today(),
                "ds_nhan_vien": phien.query(models.NhanVien).order_by(models.NhanVien.ten_nv).all(),
                "ds_ngay_nghi": phien.query(models.NgayNghi).order_by(models.NgayNghi.ngay.desc()).all(),
                "sap_xep_ngay_nghi": "ngay_moi",
                "loi_ngay_nghi": "Chưa chọn ngày nghỉ.",
            },
        )

    ngay_set = sorted(set(ngay_hop_le))
    nv_id = int(nhan_vien_id)
    for ngay in ngay_set:
        ton_tai = (
            phien.query(models.NgayNghi)
            .filter(models.NgayNghi.nhan_vien_id == nv_id)
            .filter(models.NgayNghi.ngay == ngay)
            .first()
        )
        if ton_tai:
            continue
        phien.add(models.NgayNghi(nhan_vien_id=nv_id, ngay=ngay, trang_thai="OFF", ghi_chu=None))
    phien.commit()
    return RedirectResponse(url="/", status_code=303)


@ung_dung.post("/ngay-nghi/{ngay_nghi_id}/xoa")
def xoa_ngay_nghi(
    ngay_nghi_id: int,
    redirect_to: str | None = Form(None),
    phien: Session = Depends(lay_phien_lam_viec),
):
    nn = phien.query(models.NgayNghi).filter(models.NgayNghi.id == ngay_nghi_id).first()
    if nn:
        phien.delete(nn)
        phien.commit()
    return RedirectResponse(url=redirect_to or "/ngay-nghi", status_code=303)


@ung_dung.post("/ngay-nghi/xoa-tat-ca")
def xoa_ngay_nghi_tat_ca(
    redirect_to: str | None = Form(None),
    phien: Session = Depends(lay_phien_lam_viec),
):
    phien.query(models.NgayNghi).delete()
    phien.commit()
    return RedirectResponse(url=redirect_to or "/ngay-nghi", status_code=303)


@ung_dung.get("/nhu-cau-ca")
def nhu_cau_ca(request: Request, phien: Session = Depends(lay_phien_lam_viec)):
    ds_nhu_cau = phien.query(models.NhuCauCa).order_by(models.NhuCauCa.ngay.desc()).all()
    ds_chi_nhanh = phien.query(models.ChiNhanh).order_by(models.ChiNhanh.id).all()
    ds_ca = phien.query(models.CaLam).order_by(models.CaLam.id).all()
    ds_vai_tro = phien.query(models.VaiTro).order_by(models.VaiTro.id).all()
    return giao_dien.TemplateResponse(
        "nhu_cau_ca.html",
        {
            "request": request,
            "ds_nhu_cau": ds_nhu_cau,
            "ds_chi_nhanh": ds_chi_nhanh,
            "ds_ca": ds_ca,
            "ds_vai_tro": ds_vai_tro,
        },
    )


@ung_dung.post("/nhu-cau-ca")
async def tao_nhu_cau_ca(request: Request, phien: Session = Depends(lay_phien_lam_viec)):
    form = await request.form()
    chi_nhanh_id = form.get("chi_nhanh_id")
    vai_tro_id = form.get("vai_tro_yeu_cau_id")
    crud.tao_nhu_cau_ca(
        phien,
        ngay=datetime.fromisoformat(form.get("ngay")).date(),
        chi_nhanh_id=int(chi_nhanh_id) if chi_nhanh_id else None,
        ca_id=int(form.get("ca_id")),
        so_nguoi_can=int(form.get("so_nguoi_can")),
        vai_tro_yeu_cau_id=int(vai_tro_id) if vai_tro_id else None,
        do_quan_trong=int(form.get("do_quan_trong")) if form.get("do_quan_trong") else None,
        senior_toi_thieu=int(form.get("senior_toi_thieu")) if form.get("senior_toi_thieu") else None,
    )
    return RedirectResponse(url="/nhu-cau-ca", status_code=303)


@ung_dung.post("/nhu-cau-ca/{nhu_cau_id}/xoa")
def xoa_nhu_cau_ca(nhu_cau_id: int, phien: Session = Depends(lay_phien_lam_viec)):
    nc = phien.query(models.NhuCauCa).filter(models.NhuCauCa.id == nhu_cau_id).first()
    if nc:
        phien.delete(nc)
        phien.commit()
    return RedirectResponse(url="/nhu-cau-ca", status_code=303)


@ung_dung.get("/trong-so")
def trong_so(request: Request, phien: Session = Depends(lay_phien_lam_viec)):
    ds_trong_so = phien.query(models.TrongSoUuTien).order_by(models.TrongSoUuTien.khoa).all()
    return giao_dien.TemplateResponse(
        "trong_so.html",
        {
            "request": request,
            "ds_trong_so": ds_trong_so,
            "ten_trong_so_hien_thi": ten_trong_so_hien_thi,
        },
    )


@ung_dung.post("/trong-so")
async def cap_nhat_trong_so(request: Request, phien: Session = Depends(lay_phien_lam_viec)):
    form = await request.form()
    crud.tao_hoac_cap_nhat_trong_so(
        phien,
        khoa=form.get("khoa"),
        gia_tri=int(form.get("gia_tri")),
    )
    return RedirectResponse(url="/trong-so", status_code=303)


@ung_dung.post("/trong-so/{trong_so_id}/xoa")
def xoa_trong_so(trong_so_id: int, phien: Session = Depends(lay_phien_lam_viec)):
    ts = phien.query(models.TrongSoUuTien).filter(models.TrongSoUuTien.id == trong_so_id).first()
    if ts:
        phien.delete(ts)
        phien.commit()
    return RedirectResponse(url="/trong-so", status_code=303)
